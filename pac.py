


import os
import re
import sys
import time
import pyperclip
import pandas as pd
from pywinauto import Application, keyboard
from openpyxl import load_workbook
from openpyxl.styles import Font
import secrets
from pathlib import Path
import shutil



# version comes from command line
version = sys.argv[1] if len(sys.argv) > 1 else "v1"

# base directory where script is running
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# dynamic folders based on version
PDF_FOLDER = os.path.join(BASE_DIR, version)

PAC_SKIPPED = os.path.join(BASE_DIR, version, "pac_skipped")
PAC_PROCESSED = os.path.join(BASE_DIR, version, "pac_processed")

OUTPUT_FOLDER = os.path.join(BASE_DIR, version, "pac_results")

os.makedirs(PAC_SKIPPED, exist_ok=True)
os.makedirs(PAC_PROCESSED, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# PAC application path
PAC_PATH = r"C:\Users\akash\AppData\Local\PAC\PAC.exe"

# ensure folders exist
os.makedirs(OUTPUT_FOLDER, exist_ok=True)




def get_excel_filename(pdf_name, suffix="_RESULTS.xlsx"):
    """
    Convert PDF filename to Excel filename.
    Handles both .pdf and .PDF extensions (case-insensitive).
    """
    base, _ = os.path.splitext(pdf_name)
    return base + suffix
# ---------------- HELPERS ----------------

def wait_for_results_button(pac, timeout=180):
    for _ in range(timeout):
        try:
            btn = pac.child_window(title_re="Results", control_type="Button")
            if btn.exists() and btn.is_enabled():
                return btn
        except:
            pass
        time.sleep(1)
    return None


def find_results_tree(pac, timeout=60):
    for _ in range(timeout):
        try:
            for t in pac.descendants(control_type="Tree"):
                if t.is_visible():
                    return t
        except:
            pass
        time.sleep(1)
    return None


def close_results_window(pac):
    """
    Close the Results window to prepare for next PDF
    """
    try:
        print("Closing Results window...")
        
        # Method 1: Try to find and close the Results window by title
        try:
            results_window = pac.child_window(title_re=".*Results.*", control_type="Window")
            if results_window.exists(timeout=2):
                results_window.close()
                time.sleep(1)
                print("  Results window closed via close button")
                return True
        except:
            pass
        
        # Method 2: Press ESC to close
        keyboard.send_keys("{ESC}")
        time.sleep(0.5)
        keyboard.send_keys("{ESC}")
        time.sleep(0.5)
        print("  Results window closed via ESC")
        return True
        
    except Exception as e:
        print(f"  Warning: Could not close Results window: {e}")
        # Try Alt+F4 as last resort
        try:
            keyboard.send_keys("%{F4}")
            time.sleep(0.5)
            print("  Results window closed via Alt+F4")
        except:
            pass
        return False


def refocus_pac_window(app):
    """
    Safely refocus the PAC main window
    """
    try:
        # Try to find PAC window with multiple patterns
        for title_pattern in ["PAC 2026", "PAC", ".*PAC.*"]:
            try:
                pac = app.window(title_re=title_pattern)
                if pac.exists(timeout=2):
                    pac.set_focus()
                    time.sleep(0.5)
                    return pac
            except:
                continue
        
        # If all patterns fail, just return the app top window
        return app.top_window()
    except Exception as e:
        print(f"  Warning: Could not refocus PAC window: {e}")
        return None


def type_path_safely(path):
    """
    Type file path using clipboard to avoid special char issues
    """
    pyperclip.copy(path)
    time.sleep(0.2)
    keyboard.send_keys("^v")  # Paste instead of typing
    time.sleep(0.3)


def is_end_marker(text):
    """
    Check if we've reached the end marker text
    """
    end_markers = [
        '"P" elements contain "Note" elements',
        'P" elements contain "Note" elements'
    ]
    
    for marker in end_markers:
        if marker in text:
            return True
    return False


def should_skip_subtree(text):
    """
    Check if this subtree should be skipped (not expanded)
    """
    skip_patterns = [
        "Tagged content and artifacts",
        "tagged content and artifacts"
    ]
    
    for pattern in skip_patterns:
        if pattern in text:
            return True
    return False
def check_timeout(start_time, timeout_minutes=10):
    """
    Check if extraction has exceeded timeout
    Returns True if timeout exceeded
    """
    elapsed = time.time() - start_time
    if elapsed > (timeout_minutes * 60):
        print(f"\n  TIMEOUT: Extraction exceeded {timeout_minutes} minutes")
        return True
    return False

def escape_stuck_child_properly(stuck_text, depth=0):
    """
    Properly escape from stuck child node by going UP to parent first
    """
    if depth == 0:
        print(f"  Escaping stuck section: '{stuck_text[:50]}...'")
    
    # Step 1: Go UP multiple times
    for _ in range(3 + depth):
        keyboard.send_keys("{UP}")
        time.sleep(0.05)
    
    time.sleep(0.2)
    
    # Step 2: Collapse with LEFT
    keyboard.send_keys("{LEFT}")
    time.sleep(0.15)
    
    # Step 3: Move DOWN to next sibling
    keyboard.send_keys("{DOWN}")
    time.sleep(0.15)
    
    # Step 4: Verify we escaped
    pyperclip.copy("")
    time.sleep(0.05)
    keyboard.send_keys("^c")
    time.sleep(0.1)
    new_text = pyperclip.paste().strip()
    
    if new_text and new_text != stuck_text:
        print(f"  Escaped successfully")
        return new_text
    else:
        if depth < 3:
            return escape_stuck_child_properly(stuck_text, depth + 1)
        else:
            print(f"  Using last resort escape...")
            for _ in range(10):
                keyboard.send_keys("{UP}")
                time.sleep(0.03)
            
            keyboard.send_keys("{LEFT}")
            time.sleep(0.1)
            
            for _ in range(5):
                keyboard.send_keys("{DOWN}")
                time.sleep(0.05)
            
            pyperclip.copy("")
            time.sleep(0.05)
            keyboard.send_keys("^c")
            time.sleep(0.1)
            new_text = pyperclip.paste().strip()
            
            if new_text and new_text != stuck_text:
                return new_text
            else:
                for _ in range(20):
                    keyboard.send_keys("{DOWN}")
                    time.sleep(0.03)
                return None


def expand_and_collect_until_end_marker(tree):
    """
    Collects ALL text until we find the end marker: "P" elements contain "Note" elements
    Skips expanding subtrees that match skip patterns (like "Tagged content and artifacts")
    """
    print("Starting extraction - will continue until end marker is found...")
    print('  End marker: "P" elements contain "Note" elements')
    print('  Will skip expanding: "Tagged content and artifacts"')
    tree.set_focus()
    time.sleep(0.5)
    
    # Go to top
    keyboard.send_keys("{HOME}")
    time.sleep(0.5)
    
    collected_items = []
    iteration = 0
    previous_text = ""
    consecutive_identical = 0
    escape_count = 0
    skipped_subtrees = 0
    found_end_marker = False
    start_time = time.time()
    
    while True:
        iteration += 1
        
        try:
            # First, copy current item to check what it is
            pyperclip.copy("")
            time.sleep(0.02)
            keyboard.send_keys("^c")
            time.sleep(0.08)
            current_text = pyperclip.paste().strip()
            print(current_text)
            
            if current_text:
                # CHECK FOR END MARKER FIRST
                if is_end_marker(current_text):
                    print(f"\nFOUND END MARKER: '{current_text}'")
                    collected_items.append(current_text)
                    found_end_marker = True
                    break
                
                # CHECK IF WE SHOULD SKIP THIS SUBTREE
                if should_skip_subtree(current_text):
                    print(f"  Skipping subtree: '{current_text}'")
                    skipped_subtrees += 1
                    # Collect the text but DON'T expand (no RIGHT key)
                    collected_items.append(current_text)
                    previous_text = current_text
                    # Just move DOWN without expanding
                    keyboard.send_keys("{DOWN}")
                    time.sleep(0.015)
                    continue
                
                # Normal processing - expand current node
                keyboard.send_keys("{RIGHT}")
                time.sleep(0.015)
                
                # Check if identical to previous (stuck detection)
                if current_text == previous_text:
                    consecutive_identical += 1
                    
                    # If 2 consecutive identical, escape!
                    if consecutive_identical >= 2:
                        escape_count += 1
                        
                        if escape_count % 10 == 0:
                            print(f"  Progress: {len(collected_items)} items, {escape_count} escapes, {skipped_subtrees} skipped")
                        
                        new_text = escape_stuck_child_properly(current_text)
                        
                        # Reset tracking
                        consecutive_identical = 0
                        previous_text = new_text if new_text else ""
                        
                        # Continue from new position
                        continue
                else:
                    consecutive_identical = 0
                
                # Collect the text
                collected_items.append(current_text)
                previous_text = current_text
                
                # Progress update
                if len(collected_items) % 100 == 0:
                    print(f"  Collected {len(collected_items)} items (iteration {iteration})...")
            else:
                # No text, just expand and move
                keyboard.send_keys("{RIGHT}")
                time.sleep(0.015)
            
            # Move to next item
            keyboard.send_keys("{DOWN}")
            time.sleep(0.015)
            
            if iteration % 100 == 0:  # Check every 100 iterations
                if check_timeout(start_time, timeout_minutes=10):
                    print(f"  Stopping extraction due to timeout")
                    print(f"  Collected {len(collected_items)} items before timeout")
                    break
            
            # Safety limit
            if iteration >= 100000:
                print(f"\nReached safety limit of 100,000 iterations without finding end marker")
                print(f"Last text seen: '{current_text[:100]}...'")
                break
                
        except Exception as e:
            if iteration % 500 == 0:
                print(f"  Minor error at iteration {iteration}: {e}")
            continue
    
    print(f"\nExtraction complete!")
    print(f"  Total items collected: {len(collected_items)}")
    print(f"  Total iterations: {iteration}")
    print(f"  Total escapes performed: {escape_count}")
    print(f"  Subtrees skipped: {skipped_subtrees}")
    print(f"  End marker found: {found_end_marker}")
    
    # Remove duplicates
    seen = set()
    unique_items = []
    for item in collected_items:
        if item not in seen:
            seen.add(item)
            unique_items.append(item)
    
    duplicate_count = len(collected_items) - len(unique_items)
    print(f"  Unique items: {len(unique_items)}")
    if duplicate_count > 0:
        print(f"  Duplicates removed: {duplicate_count}")
    
    return unique_items


def save_list_to_excel(data_list, output_path):
    """
    Save list to Excel
    """
    try:
        if not data_list:
            print("No data to save")
            return False
        
        data_list = [item for item in data_list if item.strip()]
        
        if not data_list:
            print("No valid data after filtering")
            return False
        
        df = pd.DataFrame({"Results": data_list})
        df.to_excel(output_path, index=False, engine='openpyxl')
        print(f"SUCCESS: Saved {len(data_list)} lines to Excel")
        print(f"File: {output_path}")
        return True
        
    except Exception as e:
        print(f"ERROR saving to Excel: {e}")
        return False
def parse_pac_results(lines):
    rows = []
    current_category = "General"
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip header rows
        if "Passed" in line and "Warning" in line and "Error" in line:
            continue
        parts = line.split()
        # Row with counts at the end
        if len(parts) >= 4 and (parts[-1].isdigit() or parts[-1] == "-"):
            rows.append({
                "Category": current_category,
                "Checkpoint": " ".join(parts[:-3]),
                "Passed": parts[-3],
                "Warning": parts[-2],
                "Error": parts[-1],
                "Message": ""
            })
        else:
            # Category or error message
            current_category = line
            rows.append({
                "Category": current_category,
                "Checkpoint": "",
                "Passed": "",
                "Warning": "",
                "Error": "",
                "Message": line
            })
    return rows

def save_structured_results_to_excel(raw_lines, output_path):
    rows = []
    parent_rows = []
    last_row_was_parent = False
    checkList = [ 'pdf syntax', "parents of structure elements", "logical structure syntax", "structural parent tree", "registry entries in type 0 fonts", "ordering entries in type 0 fonts", "supplement entries in type 0 fonts", "cid to gid mapping of type 2 cid fonts", "predefined or embedded cmaps", "wmode entry in cmap definition and cmap data", "references inside cmaps to other cmaps", "font embedding", "encoding entry in non-symbolic truetype font", "encoding of symbolic truetype fonts", "glyph names in non-symbolic truetype font", "tagged content and artifacts", "artifacts inside tagged content", "tagged content inside artifacts", "mapping of characters to unicode", "referenced external objects", "name entry in occds (optional content configuration dictionaries)", "as entry in occds (optional content configuration dictionaries)", "f and uf entries in file specifications", "correctness of language attribute", "natural language of text objects", "natural language of alternative text", "natural language of actual text", "natural language of expansion text", "natural language of bookmarks (document outline)", "natural language of contents entries in annotations", "natural language of alternate names of form fields", "use of either h or hn structure elements", "first heading level", "nesting of heading levels", "h structure elements within a structure node", "ids of note structure elements", "unique id entries in note structure elements", "trapnet annotations", "nesting of widget annotations inside a form structure elements", "nesting of link annotations inside link structure elements", "nesting of annotations in annot structure elements", "printermark annotations", "bounding boxes", "table regularity", "table header cell assignments", "document structure elements", "part structure elements", "art structure elements", "sect structure elements", "div structure elements", "blockquote structure elements", "caption structure elements", "toc structure elements", "toci structure elements", "index structure elements", "private structure elements", "h structure elements", "h1 structure elements", "h2 structure elements", "h3 structure elements", "h4 structure elements", "h5 structure elements", "h6 structure elements", "p structure elements", "l structure elements", "li structure elements", "lbl structure elements", "lbody structure elements", "table structure elements", "tr structure elements", "th structure elements", "td structure elements", "thead structure elements", "tbody structure elements", "tfoot structure elements", "span structure elements", "quote structure elements", "note structure elements", "reference structure elements", "bibentry structure elements", "code structure elements", "link structure elements", "annot structure elements", "ruby structure elements", "rb structure elements", "rt structure elements", "rp structure elements", "warichu structure elements", "wp structure elements", "wt structure elements", "figure structure elements", "formula structure elements", "form structure elements", "content is present in admissible locations", "role mapping for standard structure types", "role mapping of non-standard structure types", "circular role mapping", "alternative text for figure structure elements", "alternative text for formula structure elements", "alternate names for form fields", "alternative description for annotations", "xmp metadata", "pdf/ua identifier", "title in xmp metadata", "display of document title in window title", "tag suspects", "mark for tagged documents", "dynamic xfa form", "security settings and document access by assistive technologies", "tab order for pages with annotations", "1.2.1 audio-only and video-only (prerecorded)", "1.2.2 captions (prerecorded)", "1.2.3 audio description or media alternative (prerecorded)", "1.2.4 captions (live)", "1.2.5 audio description (prerecorded)", "1.3.2 meaningful sequence", "1.3.3 sensory characteristics", "1.3.4 orientation", "1.3.5 identify input purpose", "1.4.1 use of color", "1.4.2 audio control", "1.4.3 contrast (minimum)", "1.4.4 resize text", "1.4.5 images of text", "1.4.10 reflow", "1.4.11 non-text contrast", "1.4.12 text spacing", "1.4.13 content on hover or focus", "2.1.1 keyboard", "2.1.2 no keyboard trap", "2.1.4 character key shortcuts", "2.2.1 timing adjustable", "2.2.2 pause, stop, hide", "2.3.1 three flashes or below threshold", "2.4.1 bypass blocks", "2.4.4 link purpose (in context)", "2.4.5 multiple ways", "2.4.6 headings and labels", "2.4.7 focus visible", "2.5.1 pointer gestures", "2.5.2 pointer cancellation", "2.5.3 label in name", "2.5.4 motion actuation", "3.2.1 on focus", "3.2.2 on input", "3.2.3 consistent navigation", "3.2.4 consistent identification", "3.3.1 error identification", "3.3.2 labels or instructions", "3.3.3 error suggestion", "3.3.4 error prevention (legal, financial, data)", "4.1.3 status messages", "validity of document title", "artifacted content on page body", "tagged text consists of only whitespace", "tagged content exists outside of the page boundaries", "presence of headings", "presence of bookmarks (document outline) if there are headings", "toci elements contain link elements", "toci elements are correctly linked to headings", "validity of alternative texts", "alternative text on text elements", "completeness of link elements", "formal correctness of li elements", "completeness of table elements", "note elements are referenced", "note elements contain lbl elements", "p elements contain note elements" ]
    finalCheckList = {norm(x) for x in checkList}
    def is_parent_node(checkpoint, passed):
        return checkpoint and passed.strip().isdigit()
    for block in raw_lines:
        for line in block.split("\n"):
            line = line.strip()
            if not line or line.startswith("Checkpoint"):
                continue
            parts = line.split("\t")
            # -----------------------
            # STRUCTURED ROW
            # -----------------------
            if len(parts) == 4:
                checkpoint, passed, warning, error = parts

                if norm(checkpoint) in finalCheckList and is_parent_node(checkpoint, passed):
                    # 🔹 If previous parent group ended, add ONE blank row
                    if last_row_was_parent is False and rows:
                        rows.append({
                            "Checkpoint": "",
                            "Passed": "",
                            "Warning": "",
                            "Error": "",
                            "Message": ""
                        })
                    parent_rows.append(len(rows) + 2)  # Excel row index
                    last_row_was_parent = True
                    rows.append({
                        "Checkpoint": checkpoint,
                        "Passed": passed,
                        "Warning": warning,
                        "Error": error,
                        "Message": ""
                    })
                elif norm(checkpoint) in finalCheckList:
                    last_row_was_parent = False
                    rows.append({
                        "Checkpoint": checkpoint,
                        "Passed": passed,
                        "Warning": warning,
                        "Error": error,
                        "Message": ""
                    })
            # -----------------------
            # MESSAGE / DETAIL ROW
            # -----------------------
    # ✅ Final blank row after last parent group

    rows.append({
        "Checkpoint": "",
        "Passed": "",
        "Warning": "",
        "Error": "",
        "Message": ""
    })
    # -----------------------
    # SAVE TO EXCEL
    # -----------------------
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, engine="openpyxl")
    wb = load_workbook(output_path)
    ws = wb.active
    # Column widths
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 90
    # Bold parent rows
    bold_font = Font(bold=True)
    for r in parent_rows:
        for c in ["A", "B", "C", "D"]:
            ws[f"{c}{r}"].font = bold_font
    wb.save(output_path)
    print(f"Excel saved (row spacing after each parent group): {output_path}")
    return True

# ================= FINAL SUMMARY HELPERS =================

def derive_status(passed, warning, error):
    statuses = []
    def num(v):
        return int(v) if v.strip().isdigit() else 0
    p = num(passed)
    w = num(warning)
    e = num(error)
    if p > 0:
        statuses.append("passed")
    if w > 0:
        statuses.append("warning")
    if e > 0:
        statuses.append("failed")
    # If no values at all
    if not statuses:
        return "skipped"
    return " / ".join(statuses)
def norm(s: str) -> str:
    s = s.replace("\xa0", " ")      # NBSP -> space (safe)
    s = s.strip().lower()
    s = s.replace('"', '')          # remove quotes
    s = re.sub(r"\s+", " ", s)      # collapse whitespace
    return s

def extract_summary_from_raw(raw_lines, pdf_filename):
    checkList = [ 'pdf syntax', "parents of structure elements", "logical structure syntax", "structural parent tree", "registry entries in type 0 fonts", "ordering entries in type 0 fonts", "supplement entries in type 0 fonts", "cid to gid mapping of type 2 cid fonts", "predefined or embedded cmaps", "wmode entry in cmap definition and cmap data", "references inside cmaps to other cmaps", "font embedding", "encoding entry in non-symbolic truetype font", "encoding of symbolic truetype fonts", "glyph names in non-symbolic truetype font", "tagged content and artifacts", "artifacts inside tagged content", "tagged content inside artifacts", "mapping of characters to unicode", "referenced external objects", "name entry in occds (optional content configuration dictionaries)", "as entry in occds (optional content configuration dictionaries)", "f and uf entries in file specifications", "correctness of language attribute", "natural language of text objects", "natural language of alternative text", "natural language of actual text", "natural language of expansion text", "natural language of bookmarks (document outline)", "natural language of contents entries in annotations", "natural language of alternate names of form fields", "use of either h or hn structure elements", "first heading level", "nesting of heading levels", "h structure elements within a structure node", "ids of note structure elements", "unique id entries in note structure elements", "trapnet annotations", "nesting of widget annotations inside a form structure elements", "nesting of link annotations inside link structure elements", "nesting of annotations in annot structure elements", "printermark annotations", "bounding boxes", "table regularity", "table header cell assignments", "document structure elements", "part structure elements", "art structure elements", "sect structure elements", "div structure elements", "blockquote structure elements", "caption structure elements", "toc structure elements", "toci structure elements", "index structure elements", "private structure elements", "h structure elements", "h1 structure elements", "h2 structure elements", "h3 structure elements", "h4 structure elements", "h5 structure elements", "h6 structure elements", "p structure elements", "l structure elements", "li structure elements", "lbl structure elements", "lbody structure elements", "table structure elements", "tr structure elements", "th structure elements", "td structure elements", "thead structure elements", "tbody structure elements", "tfoot structure elements", "span structure elements", "quote structure elements", "note structure elements", "reference structure elements", "bibentry structure elements", "code structure elements", "link structure elements", "annot structure elements", "ruby structure elements", "rb structure elements", "rt structure elements", "rp structure elements", "warichu structure elements", "wp structure elements", "wt structure elements", "figure structure elements", "formula structure elements", "form structure elements", "content is present in admissible locations", "role mapping for standard structure types", "role mapping of non-standard structure types", "circular role mapping", "alternative text for figure structure elements", "alternative text for formula structure elements", "alternate names for form fields", "alternative description for annotations", "xmp metadata", "pdf/ua identifier", "title in xmp metadata", "display of document title in window title", "tag suspects", "mark for tagged documents", "dynamic xfa form", "security settings and document access by assistive technologies", "tab order for pages with annotations", "1.2.1 audio-only and video-only (prerecorded)", "1.2.2 captions (prerecorded)", "1.2.3 audio description or media alternative (prerecorded)", "1.2.4 captions (live)", "1.2.5 audio description (prerecorded)", "1.3.2 meaningful sequence", "1.3.3 sensory characteristics", "1.3.4 orientation", "1.3.5 identify input purpose", "1.4.1 use of color", "1.4.2 audio control", "1.4.3 contrast (minimum)", "1.4.4 resize text", "1.4.5 images of text", "1.4.10 reflow", "1.4.11 non-text contrast", "1.4.12 text spacing", "1.4.13 content on hover or focus", "2.1.1 keyboard", "2.1.2 no keyboard trap", "2.1.4 character key shortcuts", "2.2.1 timing adjustable", "2.2.2 pause, stop, hide", "2.3.1 three flashes or below threshold", "2.4.1 bypass blocks", "2.4.4 link purpose (in context)", "2.4.5 multiple ways", "2.4.6 headings and labels", "2.4.7 focus visible", "2.5.1 pointer gestures", "2.5.2 pointer cancellation", "2.5.3 label in name", "2.5.4 motion actuation", "3.2.1 on focus", "3.2.2 on input", "3.2.3 consistent navigation", "3.2.4 consistent identification", "3.3.1 error identification", "3.3.2 labels or instructions", "3.3.3 error suggestion", "3.3.4 error prevention (legal, financial, data)", "4.1.3 status messages", "validity of document title", "artifacted content on page body", "tagged text consists of only whitespace", "tagged content exists outside of the page boundaries", "presence of headings", "presence of bookmarks (document outline) if there are headings", "toci elements contain link elements", "toci elements are correctly linked to headings", "validity of alternative texts", "alternative text on text elements", "completeness of link elements", "formal correctness of li elements", "completeness of table elements", "note elements are referenced", "note elements contain lbl elements", "p elements contain note elements" ]
    finalCheckList = {norm(x) for x in checkList}
    
    summary = {"File Name": pdf_filename, "File ID" :secrets.token_hex(32)}
    for block in raw_lines:
        for line in block.split("\n"):
            line = line.strip()
            if not line or line.startswith("Checkpoint"):
                continue
            parts = line.split("\t")
            if len(parts) == 4:
                checkpoint, passed, warning, error = parts
                # ✅ INCLUDE PARENT + CHILD NODES
                #print("checkpoint repr:", repr(checkpoint))
                #print("checkpoint len :", len(checkpoint))
                #print("example item repr:", repr(next(iter(finalCheckList))))
                if norm(checkpoint) in finalCheckList:
                    status = derive_status(passed, warning, error)
                    summary[checkpoint] = status
    return summary
def update_final_summary_sheet(excel_path, summary_row):
    # If file does not exist → create it
    if not os.path.exists(excel_path):
        df = pd.DataFrame([summary_row])
        df.to_excel(
            excel_path,
            sheet_name="Final_Summary",
            index=False
        )
        print(f"Final summary file created: {excel_path}")
        return

    try:
        # Load exactly the existing workbook to update IN PLACE
        wb = load_workbook(excel_path)
        
        # Check if the user's specific sheet exists, else fallback to active sheet
        sheet_candidates = ["Final_Summary", "Pac Accessibility Report"]
        ws = None
        for name in sheet_candidates:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Map current column headers to their indices (1-based index in openpyxl)
        header_row_idx = 1
        col_map = {}
        for cell in ws[header_row_idx]:
            if cell.value is not None:
                col_name = str(cell.value).strip().lower()
                col_map[col_name] = cell.column

        # Make sure our new keys exist in the header if they are not there
        current_max_col = ws.max_column
        for key in summary_row.keys():
            key_lower = key.strip().lower()
            if key_lower not in col_map:
                current_max_col += 1
                col_map[key_lower] = current_max_col
                ws.cell(row=header_row_idx, column=current_max_col, value=key)

        # Look for the target File Name to update exactly that row
        target_file_name = summary_row.get("File Name", "")
        file_name_col = col_map.get("file name")
        
        target_row_idx = None
        if file_name_col:
            # Check row by row to find the matching File Name
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=file_name_col).value
                if str(cell_val).strip() == str(target_file_name).strip():
                    target_row_idx = row_idx
                    break

        # If not found, add it to the end
        if target_row_idx is None:
            target_row_idx = ws.max_row + 1

        # Now, strictly update ONLY this row with new data
        for key, value in summary_row.items():
            key_lower = key.strip().lower()
            col_idx = col_map[key_lower]
            ws.cell(row=target_row_idx, column=col_idx, value=value)

        # Optional: Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 5, 60)

        # Save changes to the same file path without altering anything else
        wb.save(excel_path)
        print(f"Final summary updated (File '{target_file_name}' overwritten exactly in place)")

    except Exception as e:
        print(f"Warning: Could not update Excel file safely in place: {e}")

def update_final_summary_sheet1(excel_path, summary_row):
    # Use perfectly identical behavior for safe in-place replace
    update_final_summary_sheet(excel_path, summary_row)

def check_for_error_dialog(app, timeout=5):
    """
    Detect if PAC error dialog appeared, close it, and return True if error occurred.
    
    The error dialog "Error – PAC 26.0.0.0" is a SEPARATE top-level window.
    
    Strategy:
      1. pywinauto findwindows + set_focus() + click_input() (REAL mouse click) - PRIMARY
      2. ctypes EnumWindows as fallback
      3. VERIFY the dialog is actually closed
      4. Refocus the PAC main window
    """
    start_time = time.time()
    print("  Checking for error dialogs...")
    
    time.sleep(5)  # Wait longer for dialog to fully appear
    
    # Words that indicate a false positive (not the real error dialog)
    FALSE_POSITIVES = ["antigravity", "visual studio", "code", "chrome", 
                       "firefox", "edge", "explorer", "notepad"]
    
    error_occurred = False
    dialog_closed = False
    error_hwnd = None
    
    elapsed = lambda: f"[{time.time() - start_time:.1f}s]"
    
    # =====================================================================
    # METHOD 1 (PRIMARY): pywinauto findwindows + set_focus + click_input
    # =====================================================================
    try:
        from pywinauto import findwindows
        from pywinauto import Application as PwApp
        print(f"    {elapsed()} Method 1: pywinauto findwindows + click_input...")
        
        search_patterns = ["Error.*PAC 26", "Error.*PAC"]
        
        for pattern in search_patterns:
            try:
                handles = findwindows.find_windows(title_re=pattern, visible_only=True)
                for handle in handles:
                    try:
                        dlg_app = PwApp(backend="uia").connect(handle=handle)
                        dlg_win = dlg_app.window(handle=handle)
                        title = dlg_win.window_text()
                        
                        # Skip false positives
                        if any(fp in title.lower() for fp in FALSE_POSITIVES):
                            continue
                        
                        print(f"    {elapsed()} Found error dialog: '{title}'")
                        error_occurred = True
                        error_hwnd = handle
                        
                        # SET FOCUS to the error dialog
                        dlg_win.set_focus()
                        time.sleep(0.5)
                        print(f"    {elapsed()} Focused on error dialog")
                        
                        # Find OK button and do a REAL mouse click
                        try:
                            ok_btn = dlg_win.child_window(title="OK", control_type="Button")
                            if ok_btn.exists(timeout=2):
                                ok_btn.set_focus()
                                time.sleep(0.2)
                                ok_btn.click_input()  # REAL mouse click
                                print(f"    {elapsed()} ✓ Clicked OK with real mouse click")
                                time.sleep(1)
                                dialog_closed = True
                                break
                        except Exception as e:
                            print(f"    {elapsed()} Could not click OK: {e}")
                        
                        # Try clicking any button
                        if not dialog_closed:
                            try:
                                for btn in dlg_win.descendants(control_type="Button"):
                                    btn_text = btn.window_text()
                                    if btn_text.upper() in ["OK", "CLOSE"]:
                                        btn.set_focus()
                                        time.sleep(0.2)
                                        btn.click_input()
                                        print(f"    {elapsed()} ✓ Clicked '{btn_text}' with real mouse click")
                                        time.sleep(1)
                                        dialog_closed = True
                                        break
                            except Exception:
                                pass
                        
                        if dialog_closed:
                            break
                    except Exception:
                        pass
                
                if dialog_closed:
                    break
            except Exception:
                pass
    except Exception as e:
        print(f"    {elapsed()} Method 1 failed: {e}")
    
    # =====================================================================
    # METHOD 2 (FALLBACK): ctypes to find HWND + pywinauto click_input
    # =====================================================================
    if not error_occurred:
        try:
            import ctypes
            from ctypes import wintypes
            print(f"    {elapsed()} Method 2: ctypes window search...")
            
            EnumWindows = ctypes.windll.user32.EnumWindows
            GetWindowTextW = ctypes.windll.user32.GetWindowTextW
            GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
            IsWindowVisible = ctypes.windll.user32.IsWindowVisible
            
            found_hwnds = []
            
            @ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            def enum_callback(hwnd, lparam):
                if IsWindowVisible(hwnd):
                    length = GetWindowTextLengthW(hwnd)
                    if length > 0:
                        buff = ctypes.create_unicode_buffer(length + 1)
                        GetWindowTextW(hwnd, buff, length + 1)
                        title = buff.value
                        t = title.lower()
                        if "error" in t and "pac" in t:
                            if not any(fp in t for fp in FALSE_POSITIVES):
                                found_hwnds.append((hwnd, title))
                return True
            
            EnumWindows(enum_callback, 0)
            
            if found_hwnds:
                error_hwnd, error_title_found = found_hwnds[0]
                print(f"    {elapsed()} Found error dialog via ctypes: '{error_title_found}' (HWND={error_hwnd})")
                error_occurred = True
                
                # Connect via pywinauto and do real click
                try:
                    from pywinauto import Application as PwApp
                    dlg_app = PwApp(backend="uia").connect(handle=error_hwnd)
                    dlg_win = dlg_app.window(handle=error_hwnd)
                    
                    dlg_win.set_focus()
                    time.sleep(0.5)
                    print(f"    {elapsed()} Focused on error dialog")
                    
                    ok_btn = dlg_win.child_window(title="OK", control_type="Button")
                    if ok_btn.exists(timeout=2):
                        ok_btn.set_focus()
                        time.sleep(0.2)
                        ok_btn.click_input()
                        print(f"    {elapsed()} ✓ Clicked OK with real mouse click")
                        time.sleep(1)
                        dialog_closed = True
                except Exception as e:
                    print(f"    {elapsed()} pywinauto click failed: {e}")
                
                # Fallback: Focus and press ENTER
                if not dialog_closed:
                    try:
                        import ctypes
                        print(f"    {elapsed()} Fallback: Focusing dialog and pressing ENTER...")
                        ctypes.windll.user32.SetForegroundWindow(error_hwnd)
                        time.sleep(0.5)
                        keyboard.send_keys("{ENTER}")
                        time.sleep(1)
                        print(f"    {elapsed()} ✓ Sent ENTER to error dialog")
                        dialog_closed = True
                    except Exception as e:
                        print(f"    {elapsed()} ENTER fallback failed: {e}")
                
                # Fallback: WM_CLOSE
                if not dialog_closed:
                    try:
                        import ctypes
                        print(f"    {elapsed()} Fallback: Sending WM_CLOSE...")
                        ctypes.windll.user32.PostMessageW(error_hwnd, 0x0010, 0, 0)
                        time.sleep(1)
                        print(f"    {elapsed()} ✓ Sent WM_CLOSE")
                        dialog_closed = True
                    except Exception as e:
                        print(f"    {elapsed()} WM_CLOSE failed: {e}")
        except Exception as e:
            print(f"    {elapsed()} Method 2 failed: {e}")
    
    # =====================================================================
    # STEP 3: VERIFY the dialog is actually closed
    # =====================================================================
    if error_occurred and error_hwnd is not None:
        try:
            import ctypes
            time.sleep(0.5)
            still_visible = ctypes.windll.user32.IsWindowVisible(error_hwnd)
            if still_visible:
                print(f"    {elapsed()} ⚠ Dialog still visible! Trying Alt+F4...")
                ctypes.windll.user32.SetForegroundWindow(error_hwnd)
                time.sleep(0.3)
                keyboard.send_keys("%{F4}")
                time.sleep(1)
                
                still_visible = ctypes.windll.user32.IsWindowVisible(error_hwnd)
                if still_visible:
                    print(f"    {elapsed()} ⚠ Still visible! Pressing ENTER 3 times...")
                    ctypes.windll.user32.SetForegroundWindow(error_hwnd)
                    time.sleep(0.3)
                    for _ in range(3):
                        keyboard.send_keys("{ENTER}")
                        time.sleep(0.5)
                    time.sleep(1)
                    
                    still_visible = ctypes.windll.user32.IsWindowVisible(error_hwnd)
                    if still_visible:
                        print(f"    {elapsed()} ⚠ Dialog STILL visible after all attempts!")
                        dialog_closed = False
                    else:
                        print(f"    {elapsed()} ✓ Dialog closed after ENTER fallback")
                        dialog_closed = True
                else:
                    print(f"    {elapsed()} ✓ Dialog closed after Alt+F4")
                    dialog_closed = True
            else:
                print(f"    {elapsed()} ✓ Verified: error dialog is closed")
        except Exception:
            pass
        
        # =====================================================================
        # STEP 4: Wait for dialog to fully close, then refocus PAC main window
        # =====================================================================
        print(f"    {elapsed()} Waiting for error dialog to fully close...")
        time.sleep(3)
        
        try:
            print(f"    {elapsed()} Refocusing PAC main window...")
            pac_win = app.window(title_re="PAC 2026")
            if pac_win.exists(timeout=5):
                pac_win.set_focus()
                time.sleep(2)
                print(f"    {elapsed()} ✓ PAC main window refocused")
        except Exception as e:
            print(f"    {elapsed()} Warning: Could not refocus PAC: {e}")
    
    # Final status
    total_time = time.time() - start_time
    if error_occurred:
        if dialog_closed:
            print(f"    ✓ Error dialog detected and closed successfully (total: {total_time:.1f}s)")
        else:
            print(f"    ⚠ Error dialog detected but could not close (total: {total_time:.1f}s)")
    else:
        print(f"    No error dialog detected - file loaded successfully ({total_time:.1f}s)")
    
    return error_occurred


def log_error_file(pdf_filename, output_folder):
    """
    Log file that had an error during PAC processing
    """
    excel_path = os.path.join(output_folder, pdf_filename.replace(".pdf", "_RESULTS.xlsx"))
    
    # Create simple Excel with just filename and error
    df = pd.DataFrame({
        "Status": ["ERROR: File could not be processed by PAC"],
        "File Name": [pdf_filename]
    })
    df.to_excel(excel_path, index=False, engine='openpyxl')
    print(f"  ERROR file logged: {excel_path}")
    
    # Also update final summary
    final_summary_path = os.path.join(output_folder, f"PAC_Final_Summary_{version}.xlsx")
    summary_row = {
        "File Name": pdf_filename,
        "Status": "ERROR - File could not be processed"
    }
    update_final_summary_sheet(final_summary_path, summary_row)

from pathlib import Path
import shutil

def move_processed_file(file_path, processed_root):
    """
    Moves a processed file to another directory.
    Creates the directory if it doesn't exist.
    """
    file_path = Path(file_path)
    processed_root = Path(processed_root)

    # Create directory if it doesn't exist
    processed_root.mkdir(parents=True, exist_ok=True)

    destination = processed_root / file_path.name

    # Copy file first (works even if PAC has it locked)
    try:
        shutil.copy2(str(file_path), str(destination))
        print(f"Copied: {file_path.name} → {destination}")
    except Exception as e:
        print(f"Warning: Could not copy {file_path.name}: {e}")
        return
    
    # Try to delete the original (may fail if PAC still has it open)
    try:
        os.remove(str(file_path))
        print(f"Deleted original: {file_path.name}")
    except PermissionError:
        print(f"Note: Original file still locked by PAC, will be left in place: {file_path.name}")
    except Exception as e:
        print(f"Warning: Could not delete original: {e}")

# ---------------- START PAC ----------------

if not os.path.isdir(PDF_FOLDER):
    print(f"Input folder not found for version '{version}': {PDF_FOLDER}")
    print("Nothing to process. Please create the folder and add PDF files.")
    raise SystemExit(0)

pdfs = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith(".pdf")]
if not pdfs:
    print(f"No PDF files found in: {PDF_FOLDER}")
    print("Nothing to process. Add PDF files and run again.")
    raise SystemExit(0)

print("Starting PAC application...")

app = Application(backend="uia").start(PAC_PATH)
pac = app.window(title_re="PAC 2026")

pac.wait("exists", timeout=90)
pac.maximize()
time.sleep(6)

print(f"Found {len(pdfs)} PDF files to process\n")

# ---------------- MAIN LOOP ----------------

for idx, pdf in enumerate(pdfs, 1):
    print(f"\n{'='*70}")
    print(f"Processing [{idx}/{len(pdfs)}]: {pdf}")
    print(f"{'='*70}")

    pdf_path = os.path.join(PDF_FOLDER, pdf)
    
    print("Step 1: Opening PDF in PAC...")
    pac = refocus_pac_window(app)
    if not pac:
        print("ERROR: Could not find PAC window - skipping this PDF")
        continue
    
    time.sleep(0.5)
    keyboard.send_keys("^o")
    time.sleep(1.5)
    
    print(f"  Loading file: {pdf_path}")
    type_path_safely(pdf_path)
    keyboard.send_keys("{ENTER}")
    time.sleep(3)
    
    print("Step 2: Waiting for analysis to complete...")
    results_btn = wait_for_results_button(pac)
    
    # AFTER analysis is complete, check for error dialog
    # The error dialog only appears AFTER PAC finishes analyzing the PDF
    if check_for_error_dialog(app, timeout=5):
        print("  Warning: Error dialog was closed, continuing to extract results...")
    
    # If Results button wasn't found (error dialog may have been blocking), try again
    if not results_btn:
        print("  Retrying: Looking for Results button after closing error dialog...")
        pac = refocus_pac_window(app)
        results_btn = wait_for_results_button(pac, timeout=30)
    
    if not results_btn:
        print("ERROR: Results button not found - skipping this PDF")
        log_error_file(pdf, OUTPUT_FOLDER)
        continue

    print("Step 3: Opening Results window...")
    results_btn.click_input()
    time.sleep(4)

    print("Step 4: Locating results tree...")
    tree = find_results_tree(pac)
    if not tree:
        print("ERROR: Results tree not found - skipping this PDF")
        close_results_window(pac)
        continue

    print("Step 5: Extracting until end marker...")
    extracted_items = expand_and_collect_until_end_marker(tree)

    print("\nStep 6: Saving screenshot...")
    out_img = os.path.join(OUTPUT_FOLDER, pdf.replace(".pdf", "_RESULTS.png"))

    print("Step 7: Saving to Excel...")
    if extracted_items:
        excel_path = os.path.join(OUTPUT_FOLDER, pdf.replace(".pdf", "_RESULTS.xlsx"))
        save_structured_results_to_excel(
            extracted_items,
            excel_path
        )
        final_summary_path = os.path.join(
            OUTPUT_FOLDER,
            f"PAC_Final_Summary_{version}.xlsx"
            )
        summary_row = extract_summary_from_raw(
            extracted_items,
            pdf
        )

        update_final_summary_sheet(
            final_summary_path,
            summary_row
        )   
        
        print("\nPreview (first 15 lines):")
        for i, line in enumerate(extracted_items[:15]):
            print(f"  {i+1}. {line[:80]}...")
        if len(extracted_items) > 15:
            print(f"  ... and {len(extracted_items) - 15} more lines")
        
        if len(extracted_items) > 20:
            print("\nLast 5 lines:")
            for i, line in enumerate(extracted_items[-5:], start=len(extracted_items)-4):
                print(f"  {i}. {line[:80]}...")
    else:
        print("WARNING: No content was extracted")

    print("\nStep 8: Closing Results window...")
    close_results_window(pac)
    time.sleep(1)
    
    print("Step 9: Preparing for next file...")
    pac = refocus_pac_window(app)
    move_processed_file(pdf_path, PAC_PROCESSED)
    time.sleep(0.5)

    print(f"\nCompleted processing: {pdf}")
    time.sleep(1)

print("\n" + "="*70)
print("All PDFs processed successfully!")
print("="*70)

try:
    pac.close()
except:
    pass

app.kill()
