import os
from dotenv import load_dotenv
import pandas as pd
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import shutil
import time
pd.options.mode.chained_assignment = None

#step1 prepvspac comparsion report
def generate_accessibility_report(pac_file, prep_file, output_file):

    print("🔄 Generating Final Accessibility Report...")

    # 🔥 YOUR FULL COLUMN MAPPING
    column_mapping = { 
        # (⚠️ Your FULL mapping stays here exactly as you pasted earlier)
        # To keep message clean, assume it is fully pasted here unchanged
  "PDF syntax": "PDFSyntax",
  "Parents of structure elements": "Parents of structure elements",
  "Logical structure syntax": "Logical structure syntax",
  "Structural parent tree": "Structural Parent Tree",
  "\"Registry\" entries in Type 0 fonts": "Registry Entries In Type 0 Fonts",
  "\"Ordering\" entries in Type 0 fonts": "Ordering Entries In Type 0 Fonts",
  "\"Supplement\" entries in Type 0 fonts": "Supplement Entries In Type 0 Fonts",
  "\"CID\" to \"GID\" mapping of Type 2 CID fonts": "CID To GID Mapping Of Type 2 CID Fonts",
  "Predefined or embedded CMaps": "PREP_NA",
  "\"WMode\" entry in CMap definition and CMap data": "'WMode' Entry In CMap Definition And CMap Data",
  "References inside CMaps to other CMaps": "PREP_NA",
  "Font embedding": "Font Embedded",
  "Encoding entry in non-symbolic TrueType font": "Encoding Entry In Non-Symbolic TrueType Font",
  "Encoding of symbolic TrueType fonts": "Encoding Of Symbolic TrueType Fonts",
  "Glyph names in non-symbolic TrueType font": "PREP_NA",
  "Tagged content and artifacts": "Tagged Content",
  "Artifacts inside tagged content": "Tagged Content",
  "Tagged content inside artifacts": "Tagged Content",
  "Mapping of characters to Unicode": "Character Encoding",
  "Referenced external objects": "Referenced external objects",
  "Name entry in OCCDs (optional content configuration dictionaries)": "Name entry in OCCD",
  "AS entry in OCCDs (optional content configuration dictionaries)": "AS entry in OCCD",
  "\"F\" and \"UF\" entries in file specifications": "F And UF Entries In File Specifications",
  "Correctness of language attribute": "Correctness of language attribute",
  "Natural language of text objects": "Natural language of text objects",
  "Natural language of alternative text": "Natural Language Alternative Text",
  "Natural language of actual text": "Natural Language Actual Text",
  "Natural language of expansion text": "PREP_NA",
  "Natural language of bookmarks (document outline)": "Natural Language Of bookmark",
  "Natural language of \"Contents\" entries in annotations": "Natural Language Of Content Entries In Annotation",
  "Natural language of alternate names of form fields": "Natural Language Of An Alternate Of Form",
  "Use of either \"H\" or \"Hn\" structure elements": "Use Of H Or Hn Structure Elements",
  "First heading level": "Appropriate Nesting",
  "Nesting of heading levels": "Appropriate Nesting",
  "\"H\" structure elements within a structure node": "H Structure Elements Within A Structure Node",
  "IDs of \"Note\" structure elements": "ID's Of 'Note' Structure Element",
  "Unique\" ID\" entries in Note structure elements": "Unique ID Entries In Note Structure Elements",
  "\"TrapNet\" annotations": "TrapNet Annotation",
  "Nesting of \"Widget\" annotations inside a \"Form\" structure elements": "Tagged Annotations",
  "Nesting of \"Link\" annotations inside \"Link\" structure elements": "Tagged Annotations",
  "Nesting of annotations in Annot structure elements": "Tagged Annotations",
  "\"PrinterMark\" annotations": "PrinterMark Annotation",
  "Bounding boxes": "PREP_NA",
  "Table regularity": "Regularity",
  "Table header cell assignments": "Headers",
  "\"Document\" structure elements": "Document Structure",
  "\"Part\" structure elements": "Part Structure",
  "\"Art\" structure elements": "Art Structure",
  "\"Sect\" structure elements": "Sect Structure",
  "\"Div\" structure elements": "Div Structure",
  "\"BlockQuote\" structure elements": "Block Quote Structure",
  "\"Caption\" structure elements": "Caption Structure",
  "\"TOC\" structure elements": "TOC Structure",
  "\"TOCI\" structure elements": "TOCI Structure",
  "\"Index\" structure elements": "Index Structure",
  "\"Private\" structure elements": "Private Structure",
  "\"H\" structure elements": "H Structure",
  "\"H1\" structure elements": "H1 Structure",
  "\"H2\" structure elements": "H2 Structure",
  "\"H3\" structure elements": "H3 Structure",
  "\"H4\" structure elements": "H4 Structure",
  "\"H5\" structure elements": "H5 Structure",
  "\"H6\" structure elements": "H6 Structure",
  "\"P\" structure elements": "P Structure",
  "\"L\" structure elements": "List Structure",
  "\"LI\" structure elements": "LI Structure",
  "\"Lbl\" structure elements": "Lbl Structure",
  "\"LBody\" structure elements": "LBody Structure",
  "\"Table\" structure elements": "Table Structure",
  "\"TR\" structure elements": "TR Structure",
  "\"TH\" structure elements": "TH Structure",
  "\"TD\" structure elements": "TD Structure",
  "\"THead\" structure elements": "THead Structure",
  "\"TBody\" structure elements": "TBody Structure",
  "\"TFoot\" structure elements": "TFoot Structure",
  "\"Span\" structure elements": "Span Structure",
  "\"Quote\" structure elements": "Quote Structure",
  "\"Note\" structure elements": "Note Structure",
  "\"Reference\" structure elements": "Reference Structure",
  "\"BibEntry\" structure elements": "Bibliography Entry Structure",
  "\"Code\" structure elements": "Code Structure",
  "\"Link\" structure elements": "Link Structure",
  "\"Annot\" structure elements": "Annotation Structure",
  "\"Ruby\" structure elements": "Ruby Structure",
  "\"RB\" structure elements": "RB Structure",
  "\"RT\" structure elements": "RT Structure",
  "\"RP\" structure elements": "RP Structure",
  "\"Warichu\" structure elements": "Warichu Structure",
  "\"WP\" structure elements": "WP Structure",
  "\"WT\" structure elements": "WT Structure",
  "\"Figure\" structure elements": "Figure Structure",
  "\"Formula\" structure elements": "Formula Structure",
  "\"Form\" structure elements": "Form Structure",
  "Content is present in admissible locations": "Content is present in an Inadmissible location in the tag tree",
  "Role mapping for standard structure types": "Role Mapping Of Standard Structure Type",
  "Role mapping of non-standard structure types": "Role Mapping Of Non-Standard Structure Type",
  "Circular role mapping": "Circular Role Mapping",
  "Alternative text for \"Figure\" structure elements": "Figure Alternate Text",
  "Alternative text for \"Formula\" structure elements": "Formula Alternate text",
  "Alternate names for form fields": "Field Descriptions",
  "Alternative description for annotations": "Link Annotation Alternate text",
  "XMP Metadata": "XMP Metadata",
  "PDF/UA identifier": "PDF/UA Identifier",
  "Title in XMP metadata": "Title In XMP Metadata",
  "Display of document title in window title": "Display Of Document Title In Window Title",
  "Tag suspects": "Tag Suspects",
  "Mark for tagged documents": "Mark For Tagged Documents",
  "Dynamic XFA form": "Dynamic XFA Form",
  "Security settings and document access by assistive technologies": "Security Settings And Document Access By Assistive",
  "Tab order for pages with annotations": "Tab Order For Pages With Annotations",
  "1.2.1 Audio-only and Video-only (Prerecorded)": "1.2.1 Audio-only and Video-only(Prerecorded)",
  "1.2.2 Captions (Prerecorded)": "1.2.2 Captions(Prerecorded)",
  "1.2.3 Audio Description or Media Alternative (Prerecorded)": "1.2.3 Audio Description or Media Alternative (Prerecorded)",
  "1.2.4 Captions (Live)": "1.2.4 Captions (Live)",
  "1.2.5 Audio Description (Prerecorded)": "1.2.5 Audio Description (Prerecorded)",
  "1.3.2 Meaningful Sequence": "1.3.2 Meaningful Sequences",
  "1.3.3 Sensory Characteristics": "1.3.3 Sensory Characteristics",
  "1.3.4 Orientation": "1.3.4 Orientation",
  "1.3.5 Identify Input Purpose": "1.3.5 Identity Input Purpose",
  "1.4.1 Use of Color": "1.4.1 Use Of Color",
  "1.4.2 Audio Control": "1.4.2 Audio Control",
  "1.4.3 Contrast (Minimum)": "1.4.3 Contrast(Minimum)",
  "1.4.4 Resize text": "1.4.4 Resize Text",
  "1.4.5 Images of Text": "1.4.5 Images Of Text",
  "1.4.10 Reflow": "1.4.10 Reflow",
  "1.4.11 Non-text Contrast": "1.4.11 Non-Text Contrast",
  "1.4.12 Text Spacing": "1.4.12 Text Spacing",
  "1.4.13 Content on Hover or Focus": "1.4.13 Content On Hover Or Focus",
  "2.1.1 Keyboard": "2.1.1 Keyboard",
  "2.1.2 No Keyboard Trap": "2.1.2 No Keyboard Trap",
  "2.1.4 Character Key Shortcuts": "2.1.3 Character Key Shortcuts",
  "2.2.1 Timing Adjustable": "2.2.1 Timing Adjustable",
  "2.2.2 Pause, Stop, Hide": "2.2.2 Pause, Stop, Hide",
  "2.3.1 Three Flashes or Below Threshold": "2.3.1 Three Flashes Or Below Threshold",
  "2.4.1 Bypass Blocks": "2.4.1 Bypass Blocks",
  "2.4.4 Link Purpose (In Context)": "2.4.4 Link Purpose (In Context)",
  "2.4.5 Multiple Ways": "2.4.5 Multiple Ways",
  "2.4.6 Headings and Labels": "2.4.6 Headings And Labels",
  "2.4.7 Focus Visible": "2.4.7 Focus Visible",
  "2.5.1 Pointer Gestures": "2.5.1 Pointer Gestures",
  "2.5.2 Pointer Cancellation": "2.5.2 Pointer Cancellation",
  "2.5.3 Label in Name":"2.5.3 Label In  Name",
  "2.5.4 Motion Actuation": "2.5.4 Motion Actuation",
  "3.2.1 On Focus": "3.2.1 On Focus",
  "3.2.2 On Input":"3.2.2 On Input",
  "3.2.3 Consistent Navigation": "3.2.3 Consistent Navigation",
  "3.2.4 Consistent Identification": "3.2.4 Consistent Identification",
  "3.3.1 Error Identification": "3.3.1 Error Identification",
  "3.3.2 Labels or Instructions": "3.3.2 Labels Or Instructions",
  "3.3.3 Error Suggestion": "3.3.3 Error Suggestion",
  "3.3.4 Error Prevention (Legal, Financial, Data)": "3.3.4 Error Prevention (Legal, Financial, Data) (AA)",
  "4.1.3 Status Messages": "4.1.3 Status Messages",
  "4.1.2 (PAC Not Available)": "4.1.2 Me, Role, Value",
  "Validity of document title": "PREP_NA",
  "Artifacted content on page body": "PREP_NA",
  "Tagged text consists of only whitespace": "PREP_NA",
  "Tagged content exists outside of the page boundaries": "PREP_NA",
  "Presence of headings": "PREP_NA",
  "Presence of bookmarks (document outline) if there are headings": "PREP_NA",
  "\"TOCI\" elements contain \"Link\" elements": "PREP_NA",
  "\"TOCI\" elements are correctly linked to headings": "PREP_NA",
  "Validity of alternative texts": "PREP_NA",
  "Alternative text on text elements": "PREP_NA",
  "Completeness of \"Link\" elements": "Completeness Of Links",
  "Formal correctness of \"LI\" elements": "PREP_NA",
  "Completeness of \"Table\" elements": "PREP_NA",
  "\"Note\" elements are referenced": "PREP_NA",
  "\"Note\" elements contain \"Lbl\" elements": "PREP_NA",
  "\"P\" elements contain \"Note\" elements": "PREP_NA",
   "accessibility permission flag": "pac_NA",
  "bookmarks": "pac_NA",
  "image only pdf": "pac_NA",
  "logical reading order": "pac_NA",
  "colour contrast": "pac_NA",
  "navigation links": "pac_NA",
  "screen flicker": "pac_NA",
  "scripts": "pac_NA",
  "timed responses": "pac_NA",
   "tagged multimedia": "pac_NA",
   "hides annotation": "pac_NA"
    }

    pac_df = pd.read_excel(pac_file)
    prep_df = pd.read_excel(prep_file)

    pac_df = pac_df.drop_duplicates(subset=["File Name"])
    prep_df = prep_df.drop_duplicates(subset=["File Name"])

    def clean_col(col):
        return (
            str(col)
            .strip()
            .lower()
            .replace(" ", "")
            .replace(".", "")
            .replace('"', "")
            .replace("'", "")
        )

    pac_df.columns = pac_df.columns.str.strip()
    prep_df.columns = prep_df.columns.str.strip()

    pac_lookup = {clean_col(c): c for c in pac_df.columns}
    prep_lookup = {clean_col(c): c for c in prep_df.columns}

    merged = pd.merge(
        pac_df,
        prep_df,
        on="File Name",
        how="inner",
        suffixes=("_pac", "_prep")
    )

    def get_merged_col(base_col, suffix):
        if base_col is None:
            return None

        suffixed = f"{base_col}_{suffix}"
        if suffixed in merged.columns:
            return suffixed

        if base_col in merged.columns:
            return base_col

        return None

    def normalize(val):
        if val is None or pd.isna(val):
            return None

        s = str(val).strip().lower()
        s = s.replace("\n", "").replace("\r", "")

        if s in ["passed", "pass", "true", "yes", "1"]:
            return "pass"

        if s in ["failed", "fail", "false", "no", "0", "passed / failed"]:
            return "fail"

        if s in ["skipped", "skip"]:
            return "skip"

        if s in ["na", "n/a", "not applicable", "", "--"]:
            return None

        return None

    final_data = {"File Name": merged["File Name"]}
    json_items = list(column_mapping.items())

    for idx, (left_col, right_col) in enumerate(json_items, start=1):

        if right_col in ["pac_NA", "PREP_NA"]:
            continue

        results = []

        if idx <= 169:
            pac_name = left_col
            prep_name = right_col
        else:
            pac_name = right_col
            prep_name = left_col

        pac_clean = clean_col(pac_name)
        prep_clean = clean_col(prep_name)

        pac_base = pac_lookup.get(pac_clean)
        prep_base = prep_lookup.get(prep_clean)

        pac_col = get_merged_col(pac_base, "pac")
        prep_col = get_merged_col(prep_base, "prep")

        for _, row in merged.iterrows():

            if pac_col is None:
                pac_val = "pac_NA"
            else:
                pac_val = normalize(row[pac_col])
                if pac_val is None:
                    pac_val = "pac_NA"

            if prep_col is None:
                prep_val = "prep_NA"
            else:
                prep_val = normalize(row[prep_col])
                if prep_val is None:
                    prep_val = "prep_NA"

            if right_col in ["Tagged Annotations", "Tagged Content", "Appropriate Nesting"]:
                # Do not evaluate True/False yet! Store the raw PAC and PREP values
                results.append({"pac": pac_val, "prep": prep_val})
            
            else:
                # Standard 1-to-1 comparison
                if pac_val == "pac_NA":
                    results.append("pac_NA")
                elif prep_val == "prep_NA":
                    results.append("prep_NA")
                else:
                    results.append(pac_val == prep_val)

        # Merge results for duplicate PREP columns
        if right_col in ["Tagged Annotations", "Tagged Content", "Appropriate Nesting"] and right_col in final_data:
            existing = final_data[right_col]
            merged_results = []
            for old_dict, new_dict in zip(existing, results):
                old_pac = old_dict["pac"]
                new_pac = new_dict["pac"]
                
                # Combine PAC values: "fail" is highest priority
                if old_pac == "fail" or new_pac == "fail":
                    combined_pac = "fail"
                elif old_pac == "pass" or new_pac == "pass":
                    combined_pac = "pass"
                elif old_pac == "skip" or new_pac == "skip":
                    combined_pac = "skip"
                else:
                    combined_pac = "pac_NA"
                
                # Prep value is constant, so we can just keep the new one
                merged_results.append({"pac": combined_pac, "prep": new_dict["prep"]})
            
            final_data[right_col] = merged_results 
        else:
            final_data[right_col] = results

    # Finalize Tagged Annotations and Tagged Content: standard compare combined PAC vs PREP
    for grouped_col in ["Tagged Annotations", "Tagged Content"]:
        if grouped_col in final_data:
            final_grouped = []
            for item in final_data[grouped_col]:
                if item["pac"] == "pac_NA":
                    final_grouped.append("pac_NA")
                elif item["prep"] == "prep_NA":
                    final_grouped.append("prep_NA")
                else:
                    final_grouped.append(item["pac"] == item["prep"])
            final_data[grouped_col] = final_grouped

    # Finalize Appropriate Nesting: priority-based comparison
    # Priority rule: fail > pass > skip  (None is treated as skip)
    # Apply this rule first to combine 2 PAC keys, then again between combined PAC and PREP
    if "Appropriate Nesting" in final_data:
        final_grouped = []
        for item in final_data["Appropriate Nesting"]:
            pac = item["pac"]
            prep = item["prep"]

            # Treat None/NA as skip for priority comparison
            if pac == "pac_NA":
                pac = "skip"
            if prep == "prep_NA":
                prep = "skip"

            # Apply priority: fail > pass > skip between combined PAC and PREP
            if pac == "fail" or prep == "fail":
                final_grouped.append(False)   # fail detected anywhere
            elif pac == "pass" or prep == "pass":
                final_grouped.append(True)    # pass detected, no fail
            else:
                final_grouped.append(False)   # both skip — no confirmation of pass

        final_data["Appropriate Nesting"] = final_grouped

    final_df = pd.DataFrame(final_data)
    final_df.to_excel(output_file, index=False)
    auto_resize(output_file)

    print("✅ Final accessibility report created successfully")
    return output_file




#step2 generating the colour report
"""
PAC vs PREP Accessibility Check Comparison
============================================
For checks where multiple PAC columns map to one PREP column
(tagged content, appropriate nesting, tagged annotations),
a GROUP VERDICT is computed before the 9-combination comparison:

  PAC group verdict rule:
    - any PAC col is failed  → group = failed
    - all PAC cols skipped   → group = skipped
    - any PAC col is passed  → group = passed  (and none failed)

  PREP group verdict rule:
    - any PREP col is failed → group = failed
    - all PREP cols passed   → group = passed
    - all PREP cols skipped  → group = skipped

This produces a single merged PAC verdict vs single PREP verdict
per file per group, giving more accurate 9-combination counts.

Usage:
    python pac_prep_comparison.py

    Or import:
        from pac_prep_comparison import generate_comparison_report
        count = generate_comparison_report("PAC.xlsx", "PREP.xlsx", "out.xlsx")
        # returns grand total of (PREP Passed & PAC Skipped)

Requirements:
    pip install pandas openpyxl
"""




# ── Mapping: PAC column name -> PREP column name ──────────────────────────────
MAPPING = {
    "PDF syntax": "pdfsyntax",
    "Parents of structure elements": "parents of structure elements",
    "Logical structure syntax": "logical structure syntax",
    "Structural parent tree": "structural parent tree",
    '"Registry" entries in Type 0 fonts': "registry entries in type 0 fonts",
    '"Ordering" entries in Type 0 fonts': "ordering entries in type 0 fonts",
    '"Supplement" entries in Type 0 fonts': "supplement entries in type 0 fonts",
    '"CID" to "GID" mapping of Type 2 CID fonts': "cid to gid mapping of type 2 cid fonts",
    '"WMode" entry in CMap definition and CMap data': "'wmode' entry in cmap definition and cmap data",
    "Font embedding": "font embedded",
    "Encoding entry in non-symbolic TrueType font": "encoding entry in non-symbolic truetype font",
    "Encoding of symbolic TrueType fonts": "encoding of symbolic truetype fonts",
    # ── group: 3 PAC cols → 1 PREP col ──
    "Tagged content and artifacts":    "tagged content",
    "Artifacts inside tagged content": "tagged content",
    "Tagged content inside artifacts": "tagged content",
    "Mapping of characters to Unicode": "character encoding",
    "Referenced external objects": "referenced external objects",
    "Name entry in OCCDs (optional content configuration dictionaries)": "name entry in occd",
    "AS entry in OCCDs (optional content configuration dictionaries)": "as entry in occd",
    '"F" and "UF" entries in file specifications': "f and uf entries in file specifications",
    "Correctness of language attribute": "correctness of language attribute",
    "Natural language of text objects": "natural language of text objects",
    "Natural language of alternative text": "natural language alternative text",
    "Natural language of actual text": "natural language actual text",
    "Natural language of bookmarks (document outline)": "natural language of bookmark",
    'Natural language of "Contents" entries in annotations': "natural language of content entries in annotation",
    "Natural language of alternate names of form fields": "natural language of an alternate of form",
    'Use of either "H" or "Hn" structure elements': "use of h or hn structure elements",
    # ── group: 2 PAC cols → 1 PREP col ──
    "First heading level":    "appropriate nesting",
    "Nesting of heading levels": "appropriate nesting",
    '"H" structure elements within a structure node': "h structure elements within a structure node",
    'IDs of "Note" structure elements': "id's of 'note' structure element",
    'Unique" ID" entries in Note structure elements': "unique id entries in note structure elements",
    '"TrapNet" annotations': "trapnet annotation",
    # ── group: 3 PAC cols → 1 PREP col ──
    'Nesting of "Widget" annotations inside a "Form" structure elements': "tagged annotations",
    'Nesting of "Link" annotations inside "Link" structure elements':     "tagged annotations",
    "Nesting of annotations in Annot structure elements":                 "tagged annotations",
    '"PrinterMark" annotations': "printermark annotation",
    "Table regularity": "regularity",
    "Table header cell assignments": "headers",
    '"Document" structure elements': "document structure",
    '"Part" structure elements': "part structure",
    '"Art" structure elements': "art structure",
    '"Sect" structure elements': "sect structure",
    '"Div" structure elements': "div structure",
    '"BlockQuote" structure elements': "block quote structure",
    '"Caption" structure elements': "caption structure",
    '"TOC" structure elements': "toc structure",
    '"TOCI" structure elements': "toci structure",
    '"Index" structure elements': "index structure",
    '"Private" structure elements': "private structure",
    '"H" structure elements': "h structure",
    '"H1" structure elements': "h1 structure",
    '"H2" structure elements': "h2 structure",
    '"H3" structure elements': "h3 structure",
    '"H4" structure elements': "h4 structure",
    '"H5" structure elements': "h5 structure",
    '"H6" structure elements': "h6 structure",
    '"P" structure elements': "p structure",
    '"L" structure elements': "list structure",
    '"LI" structure elements': "li structure",
    '"Lbl" structure elements': "lbl structure",
    '"LBody" structure elements': "lbody structure",
    '"Table" structure elements': "table structure",
    '"TR" structure elements': "tr structure",
    '"TH" structure elements': "th structure",
    '"TD" structure elements': "td structure",
    '"THead" structure elements': "thead structure",
    '"TBody" structure elements': "tbody structure",
    '"TFoot" structure elements': "tfoot structure",
    '"Span" structure elements': "span structure",
    '"Quote" structure elements': "quote structure",
    '"Note" structure elements': "note structure",
    '"Reference" structure elements': "reference structure",
    '"BibEntry" structure elements': "bibliography entry structure",
    '"Code" structure elements': "code structure",
    '"Link" structure elements': "link structure",
    '"Annot" structure elements': "annotation structure",
    '"Ruby" structure elements': "ruby structure",
    '"RB" structure elements': "rb structure",
    '"RT" structure elements': "rt structure",
    '"RP" structure elements': "rp structure",
    '"Warichu" structure elements': "warichu structure",
    '"WP" structure elements': "wp structure",
    '"WT" structure elements': "wt structure",
    '"Figure" structure elements': "figure structure",
    '"Formula" structure elements': "formula structure",
    '"Form" structure elements': "form structure",
    "Content is present in admissible locations": "content is present in an inadmissible location in the tag tree",
    "Role mapping for standard structure types": "role mapping of standard structure type",
    "Role mapping of non-standard structure types": "role mapping of non-standard structure type",
    "Circular role mapping": "circular role mapping",
    'Alternative text for "Figure" structure elements': "figure alternate text",
    'Alternative text for "Formula" structure elements': "formula alternate text",
    "Alternate names for form fields": "field descriptions",
    "Alternative description for annotations": "link annotation alternate text",
    "XMP Metadata": "xmp metadata",
    "PDF/UA identifier": "pdf/ua identifier",
    "Title in XMP metadata": "title in xmp metadata",
    "Display of document title in window title": "display of document title in window title",
    "Tag suspects": "tag suspects",
    "Mark for tagged documents": "mark for tagged documents",
    "Dynamic XFA form": "dynamic xfa form",
    "Security settings and document access by assistive technologies": "security settings and document access by assistive",
    "Tab order for pages with annotations": "tab order for pages with annotations",
    "1.2.1 Audio-only and Video-only (Prerecorded)": "1.2.1 audio-only and video-only(prerecorded)",
    "1.2.2 Captions (Prerecorded)": "1.2.2 captions(prerecorded)",
    "1.2.3 Audio Description or Media Alternative (Prerecorded)": "1.2.3 audio description or media alternative (prerecorded)",
    "1.2.4 Captions (Live)": "1.2.4 captions (live)",
    "1.2.5 Audio Description (Prerecorded)": "1.2.5 audio description (prerecorded)",
    "1.3.2 Meaningful Sequence": "1.3.2 meaningful sequences",
    "1.3.3 Sensory Characteristics": "1.3.3 sensory characteristics",
    "1.3.4 Orientation": "1.3.4 orientation",
    "1.3.5 Identify Input Purpose": "1.3.5 identity input purpose",
    "1.4.1 Use of Color": "1.4.1 use of color",
    "1.4.2 Audio Control": "1.4.2 audio control",
    "1.4.3 Contrast (Minimum)": "1.4.3 contrast(minimum)",
    "1.4.4 Resize text": "1.4.4 resize text",
    "1.4.5 Images of Text": "1.4.5 images of text",
    "1.4.10 Reflow": "1.4.10 reflow",
    "1.4.11 Non-text Contrast": "1.4.11 non-text contrast",
    "1.4.12 Text Spacing": "1.4.12 text spacing",
    "1.4.13 Content on Hover or Focus": "1.4.13 content on hover or focus",
    "2.1.1 Keyboard": "2.1.1 keyboard",
    "2.1.2 No Keyboard Trap": "2.1.2 no keyboard trap",
    "2.1.4 Character Key Shortcuts": "2.1.3 character key shortcuts",
    "2.2.1 Timing Adjustable": "2.2.1 timing adjustable",
    "2.2.2 Pause, Stop, Hide": "2.2.2 pause, stop, hide",
    "2.3.1 Three Flashes or Below Threshold": "2.3.1 three flashes or below threshold",
    "2.4.1 Bypass Blocks": "2.4.1 bypass blocks",
    "2.4.4 Link Purpose (In Context)": "2.4.4 link purpose (in context)",
    "2.4.5 Multiple Ways": "2.4.5 multiple ways",
    "2.4.6 Headings and Labels": "2.4.6 headings and labels",
    "2.4.7 Focus Visible": "2.4.7 focus visible",
    "2.5.1 Pointer Gestures": "2.5.1 pointer gestures",
    "2.5.2 Pointer Cancellation": "2.5.2 pointer cancellation",
    "2.5.3 Label in Name": "2.5.3 label in name",
    "2.5.4 Motion Actuation": "2.5.4 motion actuation",
    "3.2.1 On Focus": "3.2.1 on focus",
    "3.2.2 On Input": "3.2.2 on input",
    "3.2.3 Consistent Navigation": "3.2.3 consistent navigation",
    "3.2.4 Consistent Identification": "3.2.4 consistent identification",
    "3.3.1 Error Identification": "3.3.1 error identification",
    "3.3.2 Labels or Instructions": "3.3.2 labels or instructions",
    "3.3.3 Error Suggestion": "3.3.3 error suggestion",
    "3.3.4 Error Prevention (Legal, Financial, Data)": "3.3.4 error prevention (legal, financial, data) (aa)",
    "4.1.3 Status Messages": "4.1.3 status messages",
    "4.1.2 (PAC Not Available)": "4.1.2 me, role, value",
    'Completeness of "Link" elements': "completeness of links",
}

# ── 9 combination keys & display labels ──────────────────────────────────────
COMBO_KEYS = [
    "both_passed",             # 1
    "prep_passed_pac_failed",  # 2
    "prep_failed_pac_passed",  # 3
    "prep_skipped_pac_passed", # 4
    "prep_skipped_pac_failed", # 5
    "prep_passed_pac_skipped", # 6
    "prep_failed_pac_skipped", # 7
    "both_failed",             # 8
    "both_skipped",            # 9
]
COMBO_LABELS = [
    "1. PREP Passed\n& PAC Passed",
    "2. PREP Passed\n& PAC Failed",
    "3. PREP Failed\n& PAC Passed",
    "4. PREP Skipped\n& PAC Passed",
    "5. PREP Skipped\n& PAC Failed",
    "6. PREP Passed\n& PAC Skipped",
    "7. PREP Failed\n& PAC Skipped",
    "8. Both Failed",
    "9. Both Skipped",
]
COMBO_COLORS = {
    "both_passed":             ("C6EFCE", "375623"),
    "prep_passed_pac_failed":  ("FFEB9C", "9C6500"),
    "prep_failed_pac_passed":  ("FFEB9C", "9C6500"),
    "prep_skipped_pac_passed": ("DDEBF7", "1F4E79"),
    "prep_skipped_pac_failed": ("FCE4D6", "843C0C"),
    "prep_passed_pac_skipped": ("E2EFDA", "375623"),
    "prep_failed_pac_skipped": ("FCE4D6", "843C0C"),
    "both_failed":             ("FFC7CE", "9C0006"),
    "both_skipped":            ("F2F2F2", "595959"),
}
COMBO_DESC = {
    "both_passed":             "Both tools confirmed Passed",
    "prep_passed_pac_failed":  "PREP passed but PAC flagged failure — investigate PAC finding",
    "prep_failed_pac_passed":  "PREP failed but PAC passed — PREP found an issue PAC missed",
    "prep_skipped_pac_passed": "PREP did not evaluate; PAC confirmed Passed",
    "prep_skipped_pac_failed": "PREP did not evaluate; PAC flagged failure",
    "prep_passed_pac_skipped": "PREP confirmed Passed but PAC never evaluated — coverage gap",
    "prep_failed_pac_skipped": "PREP failed but PAC skipped — PAC missed this failure",
    "both_failed":             "Both tools agree the check Failed",
    "both_skipped":            "Neither tool evaluated this check",
}


# ── Value helpers ─────────────────────────────────────────────────────────────

def _norm(v):
    if pd.isna(v): return "na"
    return str(v).strip().lower()

def _is_failed(v):  return "failed"  in _norm(v)
def _is_passed(v):  return _norm(v).startswith("passed")
def _is_skipped(v): return "skipped" in _norm(v)


# ── Group verdict helpers ─────────────────────────────────────────────────────

def _pac_group_verdict(values: list) -> str:
    """
    Merge multiple PAC column values into one verdict for a file.

    Rules (in priority order):
      1. any value is failed  → 'failed'
      2. all values skipped   → 'skipped'
      3. any value is passed  → 'passed'
      4. otherwise            → 'other'
    """
    if any(_is_failed(v) for v in values):  return "failed"
    if all(_is_skipped(v) for v in values): return "skipped"
    if any(_is_passed(v) for v in values):  return "passed"
    return "other"


def _combo(rv_verdict: str, pv_verdict: str) -> str:
    """Map (PREP verdict, PAC verdict) → one of the 9 combo keys."""
    rp = rv_verdict   # PREP
    pp = pv_verdict   # PAC
    if   rp == "passed"  and pp == "passed":  return "both_passed"
    elif rp == "passed"  and pp == "failed":  return "prep_passed_pac_failed"
    elif rp == "failed"  and pp == "passed":  return "prep_failed_pac_passed"
    elif rp == "skipped" and pp == "passed":  return "prep_skipped_pac_passed"
    elif rp == "skipped" and pp == "failed":  return "prep_skipped_pac_failed"
    elif rp == "passed"  and pp == "skipped": return "prep_passed_pac_skipped"
    elif rp == "failed"  and pp == "skipped": return "prep_failed_pac_skipped"
    elif rp == "failed"  and pp == "failed":  return "both_failed"
    elif rp == "skipped" and pp == "skipped": return "both_skipped"
    return "other"


def _na_row(display_name, prep_col, note):
    row = {k: "N/A" for k in COMBO_KEYS}
    row.update({"pac_check": display_name, "prep_check": prep_col, "note": note})
    return row


# ── Border helpers ────────────────────────────────────────────────────────────

def _thin():
    return Border(
        left=Side(style="thin",   color="BFBFBF"),
        right=Side(style="thin",  color="BFBFBF"),
        top=Side(style="thin",    color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

def _thick_top():
    return Border(
        left=Side(style="thin",   color="BFBFBF"),
        right=Side(style="thin",  color="BFBFBF"),
        top=Side(style="medium",  color="1F4E79"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


# ── Core function ─────────────────────────────────────────────────────────────

def generate_comparison_report(pac_file: str,
                                prep_file: str,
                                output_file: str,
                                save_excel: bool = True) -> int:
    """
    Compare PAC and PREP accessibility check results and write an Excel report.

    For checks where multiple PAC columns map to one PREP column, a group
    verdict is computed (any-failed > all-skipped > any-passed) so each
    file contributes exactly ONE combination result per logical check group.

    Parameters
    ----------
    pac_file    : Path to PAC_Final_Summary .xlsx
    prep_file   : Path to Prep_Accessibility_Report .xlsx
    output_file : Destination path for the output .xlsx
    save_excel  : Set False to skip writing the file (just returns the count)

    Returns
    -------
    int : Grand total of (PREP Passed & PAC Skipped) across all checks & files
    """

    # ── Load & normalise ──────────────────────────────────────────────────────
    pac  = pd.read_excel(pac_file)
    prep = pd.read_excel(prep_file)

    pac["File Name"]  = pac["File Name"].str.strip().str.lower()
    prep["File Name"] = prep["File Name"].str.strip().str.lower()
    prep = prep.rename(columns={
        c: (c.strip().lower() if c != "File Name" else "File Name")
        for c in prep.columns
    })

    prep_cols    = set(prep.columns)
    common_files = set(pac["File Name"]) & set(prep["File Name"])
    total_files  = len(common_files)

    pac_c  = pac[pac["File Name"].isin(common_files)].set_index("File Name")
    prep_c = prep[prep["File Name"].isin(common_files)].set_index("File Name")

    # ── Build groups: prep_col -> [pac_col, ...] in order ────────────────────
    # One entry per unique PREP column. Each group is processed ONCE,
    # producing a single row in the output.
    prep_to_pac_group: dict[str, list[str]] = {}
    for pac_col, prep_col in MAPPING.items():
        prep_to_pac_group.setdefault(prep_col, []).append(pac_col)

    # ── Per-check computation ─────────────────────────────────────────────────
    results = []
    grand   = {k: 0 for k in COMBO_KEYS}

    for pac_col, prep_col in MAPPING.items():

        pac_group = prep_to_pac_group[prep_col]   # all PAC cols for this PREP col
        is_group_primary = (pac_col == pac_group[0])  # only first PAC col triggers row

        # ── Skip alias rows entirely — handled inside primary row ─────────────
        if not is_group_primary:
            continue

        # ── Guard: all required columns must exist ────────────────────────────
        missing_pac  = [c for c in pac_group if c not in pac_c.columns]
        missing_prep = prep_col not in prep_cols

        if missing_pac:
            results.append(_na_row(
                " + ".join(pac_group), prep_col,
                f"PAC column(s) not found: {missing_pac}"
            ))
            continue
        if missing_prep:
            results.append(_na_row(
                " + ".join(pac_group), prep_col,
                "PREP column not found"
            ))
            continue

        prep_s  = prep_c[prep_col]
        r       = {k: 0 for k in COMBO_KEYS}

        for fname in common_files:
            # ── PREP verdict (single column) ──────────────────────────────────
            rv = prep_s.get(fname, None)
            if   _is_failed(rv):  rv_verdict = "failed"
            elif _is_passed(rv):  rv_verdict = "passed"
            elif _is_skipped(rv): rv_verdict = "skipped"
            else:                 rv_verdict = "other"

            # ── PAC group verdict (merge all PAC cols in group) ───────────────
            pac_vals   = [pac_c[c].get(fname, None) for c in pac_group]
            pv_verdict = _pac_group_verdict(pac_vals)

            # ── Map to one of 9 combos ────────────────────────────────────────
            key = _combo(rv_verdict, pv_verdict)
            if key != "other":
                r[key] += 1

        # ── Display name: join PAC cols if grouped ────────────────────────────
        if len(pac_group) == 1:
            display_name = pac_group[0]
            group_note   = ""
        else:
            display_name = pac_group[0]
            group_note   = (
                f"GROUP ({len(pac_group)} PAC cols): "
                + " | ".join(pac_group)
                + " — PAC verdict = any-failed > all-skipped > any-passed"
            )

        row_data = {
            "pac_check":  display_name,
            "prep_check": prep_col,
            "note":       group_note,
        }
        for k in COMBO_KEYS:
            row_data[k] = r[k]
        results.append(row_data)

        for k in COMBO_KEYS:
            grand[k] += r[k]

    # ── Console summary ───────────────────────────────────────────────────────
    labels_short = [
        "1. PREP Passed  & PAC Passed ",
        "2. PREP Passed  & PAC Failed ",
        "3. PREP Failed  & PAC Passed ",
        "4. PREP Skipped & PAC Passed ",
        "5. PREP Skipped & PAC Failed ",
        "6. PREP Passed  & PAC Skipped",
        "7. PREP Failed  & PAC Skipped",
        "8. Both Failed               ",
        "9. Both Skipped              ",
    ]
    print(f"\n{'='*62}")
    print(f"  PAC vs PREP — 9-Combination Summary (Group Verdict Logic)")
    print(f"{'='*62}")
    print(f"  PAC  total files          : {len(pac['File Name'].unique())}")
    print(f"  PREP total files          : {len(prep['File Name'].unique())}")
    print(f"  Common files compared     : {total_files}")
    print(f"  Total check rows          : {len(results)}")
    print(f"{'─'*62}")
    for lbl, k in zip(labels_short, COMBO_KEYS):
        print(f"  {lbl} : {grand[k]}")
    print(f"{'='*62}\n")

    if not save_excel:
        return grand

    # ── Build Excel ───────────────────────────────────────────────────────────
    HDR_BG   = "1F4E79"
    SUBHDR   = "2E75B6"
    TOTAL_BG = "1A3A5C"
    WHITE_FG = "FFFFFF"
    ALT_ROW  = "DEEAF1"
    WHITE    = "FFFFFF"
    GROUP_BG = "EEE8F5"   # light purple — highlight grouped rows
    THIN     = _thin()
    THICK    = _thick_top()

    wb = Workbook()

    # =========================================================================
    # Sheet 1 — Check Comparison
    # =========================================================================
    ws = wb.active
    ws.title = "Check Comparison"

    COL_HEADERS = ["PAC Check Name", "PREP Check Name"] + COMBO_LABELS + ["Note"]
    NUM_COLS    = len(COL_HEADERS)
    last_letter = chr(ord("A") + NUM_COLS - 1)

    # Row 1: title banner
    ws.merge_cells(f"A1:{last_letter}1")
    tc = ws["A1"]
    tc.value = (
        f"Accessibility Check Comparison: PAC vs PREP  |  {total_files} common files  |  "
        f"PREP Passed & PAC Skipped (grand total): {grand['prep_passed_pac_skipped']}"
    )
    tc.font      = Font(name="Arial", bold=True, size=12, color=WHITE_FG)
    tc.fill      = PatternFill("solid", fgColor=HDR_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: column headers
    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        if 3 <= ci <= NUM_COLS - 1:
            key    = COMBO_KEYS[ci - 3]
            bg, fg = COMBO_COLORS[key]
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Arial", bold=True, size=9, color=fg)
        else:
            c.fill = PatternFill("solid", fgColor=SUBHDR)
            c.font = Font(name="Arial", bold=True, size=9, color=WHITE_FG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[2].height = 52

    # Data rows
    for i, row in enumerate(results):
        r        = i + 3
        is_group = bool(row["note"] and "GROUP" in str(row["note"]))
        bg       = GROUP_BG if is_group else (ALT_ROW if i % 2 == 0 else WHITE)

        vals = (
            [row["pac_check"], row["prep_check"]]
            + [row[k] for k in COMBO_KEYS]
            + [row["note"]]
        )
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border    = THIN
            cell.font      = Font(name="Arial", size=9,
                                  bold=is_group and ci in [1, 2],
                                  color="000000")
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(ci in [1, 2, NUM_COLS]),
                horizontal="center" if ci not in [1, 2, NUM_COLS] else "left",
            )
            if 3 <= ci <= NUM_COLS - 1 and isinstance(val, int) and val > 0:
                key = COMBO_KEYS[ci - 3]
                cell.fill = PatternFill("solid", fgColor=COMBO_COLORS[key][0])
            elif ci == NUM_COLS and val:
                cell.fill = PatternFill("solid", fgColor="E8E0F0")
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[r].height = 22 if is_group else 18

    # Grand Total row
    total_row  = len(results) + 3
    total_vals = ["GRAND TOTAL", ""] + [grand[k] for k in COMBO_KEYS] + [""]
    for ci, val in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row, column=ci, value=val)
        cell.font      = Font(name="Arial", bold=True, size=10, color=WHITE_FG)
        cell.fill      = PatternFill("solid", fgColor=TOTAL_BG)
        cell.alignment = Alignment(
            horizontal="left" if ci <= 2 else "center", vertical="center"
        )
        cell.border = THICK
    ws.row_dimensions[total_row].height = 22

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 42
    for ci in range(3, NUM_COLS):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = 14
    ws.column_dimensions[chr(ord("A") + NUM_COLS - 1)].width = 30
    ws.freeze_panes = "C3"

    # =========================================================================
    # Sheet 2 — Grand Total Summary
    # =========================================================================
    ws2 = wb.create_sheet("Grand Total Summary")

    summary_rows = [
        ("Metric", "Grand Total", "Description"),
        ("Files in PAC",  len(pac["File Name"].unique()),
         "Total distinct files in PAC Final Summary"),
        ("Files in PREP", len(prep["File Name"].unique()),
         "Total distinct files in PREP Accessibility Report"),
        ("Common Files Compared", total_files,
         "Files present in both tools (used for all comparisons)"),
        ("Check Rows in Report",  len(results),
         "Unique logical checks after grouping multi-mapped PAC cols"),
        ("", "", ""),
    ] + [
        (lbl.replace("\n", " "), grand[k], COMBO_DESC[k])
        for lbl, k in zip(COMBO_LABELS, COMBO_KEYS)
    ]

    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value      = "Grand Total Summary — PAC vs PREP (9 Combinations, Group Verdict Logic)"
    t2.font       = Font(name="Arial", bold=True, size=13, color=WHITE_FG)
    t2.fill       = PatternFill("solid", fgColor=HDR_BG)
    t2.alignment  = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    for ci, hdr in enumerate(["Metric", "Grand Total", "Description"], 1):
        c = ws2.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE_FG)
        c.fill      = PatternFill("solid", fgColor=SUBHDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws2.row_dimensions[2].height = 22

    for i, (label, count, desc) in enumerate(summary_rows, 3):
        is_blank  = label == ""
        combo_idx = i - 7
        if 0 <= combo_idx < len(COMBO_KEYS):
            row_bg = COMBO_COLORS[COMBO_KEYS[combo_idx]][0]
        else:
            row_bg = ALT_ROW if i % 2 == 0 else WHITE

        for ci, val in enumerate([label, "" if is_blank else count, desc], 1):
            c = ws2.cell(row=i, column=ci, value=val)
            c.font      = Font(name="Arial", size=10, bold=(ci == 1 and not is_blank))
            c.alignment = Alignment(vertical="center", wrap_text=True,
                                    horizontal="center" if ci == 2 else "left")
            c.border    = THIN
            c.fill      = PatternFill("solid", fgColor=WHITE if is_blank else row_bg)
        ws2.row_dimensions[i].height = 20

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 60

    # =========================================================================
    # Sheet 3 — Group Verdict Logic Explained
    # =========================================================================
    ws3 = wb.create_sheet("Group Verdict Logic")

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value      = "Group Verdict Logic — How Multi-Mapped Checks Are Handled"
    t3.font       = Font(name="Arial", bold=True, size=13, color=WHITE_FG)
    t3.fill       = PatternFill("solid", fgColor=HDR_BG)
    t3.alignment  = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    for ci, hdr in enumerate(["Group / Check", "PAC Cols in Group", "Rule Applied"], 1):
        c = ws3.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE_FG)
        c.fill      = PatternFill("solid", fgColor=SUBHDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws3.row_dimensions[2].height = 22

    group_rows = [
        (
            "tagged content",
            "1. Tagged content and artifacts\n2. Artifacts inside tagged content\n3. Tagged content inside artifacts",
            "PAC verdict = failed if ANY of the 3 cols is failed\nPAC verdict = skipped if ALL 3 cols are skipped\nPAC verdict = passed if ANY col is passed (and none failed)"
        ),
        (
            "appropriate nesting",
            "1. First heading level\n2. Nesting of heading levels",
            "PAC verdict = failed if ANY of the 2 cols is failed\nPAC verdict = skipped if BOTH cols are skipped\nPAC verdict = passed if ANY col is passed (and none failed)"
        ),
        (
            "tagged annotations",
            "1. Nesting of 'Widget' annotations inside 'Form'\n2. Nesting of 'Link' annotations inside 'Link'\n3. Nesting of annotations in Annot structure elements",
            "PAC verdict = failed if ANY of the 3 cols is failed\nPAC verdict = skipped if ALL 3 cols are skipped\nPAC verdict = passed if ANY col is passed (and none failed)"
        ),
        (
            "All other checks\n(1 PAC col → 1 PREP col)",
            "Single PAC column",
            "No grouping needed — PAC value used directly as verdict"
        ),
    ]

    section_note = [
        ("PAC Group Verdict Priority", ""),
        ("1st priority", "any failed  → group verdict = failed"),
        ("2nd priority", "all skipped → group verdict = skipped"),
        ("3rd priority", "any passed  → group verdict = passed"),
        ("", ""),
        ("Why this order?", "A single failure anywhere in the group means the check is not fully passing. "
         "Only if every PAC col is skipped (not evaluated) do we consider the group skipped. "
         "If mixed passed/skipped with no failures, the group is considered passed."),
    ]

    r = 3
    for group_name, pac_cols_text, rule_text in group_rows:
        is_single = "Single" in pac_cols_text
        bg = GROUP_BG if not is_single else (ALT_ROW if r % 2 == 0 else WHITE)
        for ci, val in enumerate([group_name, pac_cols_text, rule_text], 1):
            c = ws3.cell(row=r, column=ci, value=val)
            c.font      = Font(name="Arial", size=9, bold=(ci == 1))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = THIN
            c.fill      = PatternFill("solid", fgColor=bg)
        ws3.row_dimensions[r].height = 52 if not is_single else 18
        r += 1

    r += 1  # blank gap
    ws3.merge_cells(f"A{r}:B{r}")
    ws3.cell(row=r, column=1, value="PAC Group Verdict Priority Rules").font = Font(
        name="Arial", bold=True, size=10, color=WHITE_FG)
    ws3.cell(row=r, column=1).fill      = PatternFill("solid", fgColor=SUBHDR)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[r].height = 22
    r += 1

    for label, desc in section_note[1:]:
        for ci, val in enumerate([label, desc], 1):
            c = ws3.cell(row=r, column=ci, value=val)
            c.font      = Font(name="Arial", size=9, bold=(ci == 1 and label))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = THIN
            c.fill      = PatternFill("solid", fgColor=ALT_ROW if r % 2 == 0 else WHITE)
        ws3.row_dimensions[r].height = 18 if label != "Why this order?" else 36
        r += 1

    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 44
    ws3.column_dimensions["C"].width = 60

    # =========================================================================
    # Sheet 4 — Legend
    # =========================================================================
    ws4 = wb.create_sheet("Legend")

    legend_items = (
        [("Column", "Meaning"),
         ("PAC Check Name",  "Check name (for grouped checks, shows the first/primary PAC col name)"),
         ("PREP Check Name", "Mapped check column in PREP Accessibility Report"),
         ("Note",            "For grouped checks: lists all PAC cols in the group and rule applied"),
         ("", ""),
         ("9 Combination Columns", "")]
        + [(lbl.replace("\n", " "),  COMBO_DESC[k])
           for lbl, k in zip(COMBO_LABELS, COMBO_KEYS)]
        + [("", ""),
           ("Light purple rows",
            "Grouped checks (multiple PAC cols merged into one verdict via group logic)"),
           ("GRAND TOTAL row (dark blue)",
            "Sum of each combination column across all check rows.")]
    )

    ws4.merge_cells("A1:B1")
    t4 = ws4["A1"]
    t4.value      = "Legend"
    t4.font       = Font(name="Arial", bold=True, size=12, color=WHITE_FG)
    t4.fill       = PatternFill("solid", fgColor=HDR_BG)
    t4.alignment  = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    for ci, hdr in enumerate(["Column / Colour", "Meaning"], 1):
        c = ws4.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE_FG)
        c.fill      = PatternFill("solid", fgColor=SUBHDR)
        c.border    = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[2].height = 20

    for i, (col, meaning) in enumerate(legend_items, 3):
        combo_idx = i - 9
        if 0 <= combo_idx < len(COMBO_KEYS):
            row_bg = COMBO_COLORS[COMBO_KEYS[combo_idx]][0]
        else:
            row_bg = ALT_ROW if i % 2 == 0 else WHITE
        for ci, val in enumerate([col, meaning], 1):
            c = ws4.cell(row=i, column=ci, value=val)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = THIN
            c.fill      = PatternFill("solid", fgColor=WHITE if col == "" else row_bg)
        ws4.row_dimensions[i].height = 18

    ws4.column_dimensions["A"].width = 46
    ws4.column_dimensions["B"].width = 72

    # ── Save & return ─────────────────────────────────────────────────────────
    wb.save(output_file)
    print(f"Saved -> {output_file}\n")
    return grand


# ── Entry point ───────────────────────────────────────────────────────────────


    
   


load_dotenv()

version = sys.argv[1] if len(sys.argv) > 1 else "v1"
CURRENT_VERSION = version





# =====================================================
# AUTO RESIZE FUNCTION
# =====================================================
def auto_resize(file_path):
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(file_path)







# =====================================================
# STEP 3 GENERATE STATUS FILE
# =====================================================
def generate_status_file(final_report_file, status_file):

    print("🔄 Creating Status File...")

    final_df = pd.read_excel(final_report_file).copy()

    def calculate_status(row):
        values = list(row.values)

        if False in values:
            return False
        elif True in values:
            return True
        elif "prep_NA" in values:
            return "Prep_NA"
        else:
            return "Pac_NA"

    final_df["Status"] = final_df.iloc[:, 1:].apply(
        calculate_status,
        axis=1
    )

    status_df = final_df[["File Name", "Status"]]
    status_df.to_excel(status_file, index=False)

    auto_resize(status_file)

    print("✅ Status file created successfully")
    return status_file


# =====================================================
# STEP 4 GENERATE VERSION SUMMARY
# =====================================================
def generate_version_summary(status_file, version_summary_file):

    print("🔄 Creating Version Summary...")

    status_df = pd.read_excel(status_file)

    status_df["Version"] = CURRENT_VERSION
    status_df["Date"] = date.today().strftime("%d %B %Y")
    status_df["Server_Name"] = "Windows Server"

    summary = status_df.groupby(
        ["Version", "Date", "Server_Name"]
    ).agg(
        Total_Files=("Status", "count"),
        Total_True=("Status", lambda x: (x == True).sum()),
        Total_False=("Status", lambda x: (x == False).sum()),
        Total_Prep_NA=("Status", lambda x: (x == "Prep_NA").sum()),
        Total_Pac_NA=("Status", lambda x: (x == "Pac_NA").sum())
    ).reset_index()

    # Check if file exists and append instead of overwrite
    import os
    if os.path.exists(version_summary_file):
        existing_df = pd.read_excel(version_summary_file)
        summary = pd.concat([existing_df, summary], ignore_index=True)

    summary.to_excel(version_summary_file, index=False)

    auto_resize(version_summary_file)

    print("✅ Version summary created successfully")
    return version_summary_file




# =====================================================
# STEP 5️ COLUMN-WISE SUMMARY
# =====================================================
def generate_final_column_summary(final_report_file):

    print("🔄 Generating Column Summary...")

    output_summary_file = os.path.join(
        os.path.dirname(final_report_file),
        f"final_summary_{version}.xlsx"
    )

    excel_file = pd.ExcelFile(final_report_file)

    with pd.ExcelWriter(output_summary_file, engine="openpyxl") as writer:
        for sheet in excel_file.sheet_names:
            df = pd.read_excel(final_report_file, sheet_name=sheet)

            summary_rows = []

            for col in df.columns[1:]:
                counts = df[col].value_counts()

                summary_rows.append({
                    "column_name": col,
                    "true_count": int(counts.get(True, 0)),
                    "false_count": int(counts.get(False, 0)),
                    "prep_na_count": int(counts.get("prep_NA", 0)),
                    "pac_na_count": int(counts.get("pac_NA", 0)),
                })

            pd.DataFrame(summary_rows).to_excel(
                writer,
                sheet_name=f"{sheet[:25]}_summary",
                index=False
            )

    auto_resize(output_summary_file)

    print("✅ Column summary generated successfully")

    return output_summary_file

#step 6 pac_prep_splitup
def generate_pac_prep_report (pac_file_path, prep_file_path, output_path) :
    # ======================================================
    # 🔥 FULL JSON MAPPING (INSIDE FUNCTION)
    # ======================================================
    check_mapping = {
     
   "PDF syntax": "PDFSyntax",
  "Parents of structure elements": "Parents of structure elements",
  "Logical structure syntax": "Logical structure syntax",
  "Structural parent tree": "Structural Parent Tree",
  "\"Registry\" entries in Type 0 fonts": "Registry Entries In Type 0 Fonts",
  "\"Ordering\" entries in Type 0 fonts": "Ordering Entries In Type 0 Fonts",
  "\"Supplement\" entries in Type 0 fonts": "Supplement Entries In Type 0 Fonts",
  "\"CID\" to \"GID\" mapping of Type 2 CID fonts": "CID To GID Mapping Of Type 2 CID Fonts",
  "Predefined or embedded CMaps": "PREP_NA",
  "\"WMode\" entry in CMap definition and CMap data": "'WMode' Entry In CMap Definition And CMap Data",
  "References inside CMaps to other CMaps": "PREP_NA",
  "Font embedding": "Font Embedded",
  "Encoding entry in non-symbolic TrueType font": "Encoding Entry In Non-Symbolic TrueType Font",
  "Encoding of symbolic TrueType fonts": "Encoding Of Symbolic TrueType Fonts",
  "Glyph names in non-symbolic TrueType font": "PREP_NA",
  "Tagged content and artifacts": "Tagged Content",
  "Artifacts inside tagged content": "Tagged Content",
  "Tagged content inside artifacts": "Tagged Content",
  "Mapping of characters to Unicode": "Character Encoding",
  "Referenced external objects": "Referenced external objects",
  "Name entry in OCCDs (optional content configuration dictionaries)": "Name entry in OCCD",
  "AS entry in OCCDs (optional content configuration dictionaries)": "AS entry in OCCD",
  "\"F\" and \"UF\" entries in file specifications": "F And UF Entries In File Specifications",
  "Correctness of language attribute": "Correctness of language attribute",
  "Natural language of text objects": "Natural language of text objects",
  "Natural language of alternative text": "Natural Language Alternative Text",
  "Natural language of actual text": "Natural Language Actual Text",
  "Natural language of expansion text": "PREP_NA",
  "Natural language of bookmarks (document outline)": "Natural Language Of bookmark",
  "Natural language of \"Contents\" entries in annotations": "Natural Language Of Content Entries In Annotation",
  "Natural language of alternate names of form fields": "Natural Language Of An Alternate Of Form",
  "Use of either \"H\" or \"Hn\" structure elements": "Use Of H Or Hn Structure Elements",
  "First heading level": "Appropriate Nesting",
  "Nesting of heading levels": "Appropriate Nesting",
  "\"H\" structure elements within a structure node": "H Structure Elements Within A Structure Node",
  "IDs of \"Note\" structure elements": "ID's Of 'Note' Structure Element",
  "Unique\" ID\" entries in Note structure elements": "Unique ID Entries In Note Structure Elements",
  "\"TrapNet\" annotations": "TrapNet Annotation",
  "Nesting of \"Widget\" annotations inside a \"Form\" structure elements": "Tagged Annotations",
  "Nesting of \"Link\" annotations inside \"Link\" structure elements": "Tagged Annotations",
  "Nesting of annotations in Annot structure elements": "Tagged Annotations",
  "\"PrinterMark\" annotations": "PrinterMark Annotation",
  "Bounding boxes": "PREP_NA",
  "Table regularity": "Regularity",
  "Table header cell assignments": "Headers",
  "\"Document\" structure elements": "Document Structure",
  "\"Part\" structure elements": "Part Structure",
  "\"Art\" structure elements": "Art Structure",
  "\"Sect\" structure elements": "Sect Structure",
  "\"Div\" structure elements": "Div Structure",
  "\"BlockQuote\" structure elements": "Block Quote Structure",
  "\"Caption\" structure elements": "Caption Structure",
  "\"TOC\" structure elements": "TOC Structure",
  "\"TOCI\" structure elements": "TOCI Structure",
  "\"Index\" structure elements": "Index Structure",
  "\"Private\" structure elements": "Private Structure",
  "\"H\" structure elements": "H Structure",
  "\"H1\" structure elements": "H1 Structure",
  "\"H2\" structure elements": "H2 Structure",
  "\"H3\" structure elements": "H3 Structure",
  "\"H4\" structure elements": "H4 Structure",
  "\"H5\" structure elements": "H5 Structure",
  "\"H6\" structure elements": "H6 Structure",
  "\"P\" structure elements": "P Structure",
  "\"L\" structure elements": "List Structure",
  "\"LI\" structure elements": "LI Structure",
  "\"Lbl\" structure elements": "Lbl Structure",
  "\"LBody\" structure elements": "LBody Structure",
  "\"Table\" structure elements": "Table Structure",
  "\"TR\" structure elements": "TR Structure",
  "\"TH\" structure elements": "TH Structure",
  "\"TD\" structure elements": "TD Structure",
  "\"THead\" structure elements": "THead Structure",
  "\"TBody\" structure elements": "TBody Structure",
  "\"TFoot\" structure elements": "TFoot Structure",
  "\"Span\" structure elements": "Span Structure",
  "\"Quote\" structure elements": "Quote Structure",
  "\"Note\" structure elements": "Note Structure",
  "\"Reference\" structure elements": "Reference Structure",
  "\"BibEntry\" structure elements": "Bibliography Entry Structure",
  "\"Code\" structure elements": "Code Structure",
  "\"Link\" structure elements": "Link Structure",
  "\"Annot\" structure elements": "Annotation Structure",
  "\"Ruby\" structure elements": "Ruby Structure",
  "\"RB\" structure elements": "RB Structure",
  "\"RT\" structure elements": "RT Structure",
  "\"RP\" structure elements": "RP Structure",
  "\"Warichu\" structure elements": "Warichu Structure",
  "\"WP\" structure elements": "WP Structure",
  "\"WT\" structure elements": "WT Structure",
  "\"Figure\" structure elements": "Figure Structure",
  "\"Formula\" structure elements": "Formula Structure",
  "\"Form\" structure elements": "Form Structure",
  "Content is present in admissible locations": "Content is present in an Inadmissible location in the tag tree",
  "Role mapping for standard structure types": "Role Mapping Of Standard Structure Type",
  "Role mapping of non-standard structure types": "Role Mapping Of Non-Standard Structure Type",
  "Circular role mapping": "Circular Role Mapping",
  "Alternative text for \"Figure\" structure elements": "Figure Alternate Text",
  "Alternative text for \"Formula\" structure elements": "Formula Alternate text",
  "Alternate names for form fields": "Field Descriptions",
  "Alternative description for annotations": "Link Annotation Alternate text",
  "XMP Metadata": "XMP Metadata",
  "PDF/UA identifier": "PDF/UA Identifier",
  "Title in XMP metadata": "Title In XMP Metadata",
  "Display of document title in window title": "Display Of Document Title In Window Title",
  "Tag suspects": "Tag Suspects",
  "Mark for tagged documents": "Mark For Tagged Documents",
  "Dynamic XFA form": "Dynamic XFA Form",
  "Security settings and document access by assistive technologies": "Security Settings And Document Access By Assistive",
  "Tab order for pages with annotations": "Tab Order For Pages With Annotations",
  "1.2.1 Audio-only and Video-only (Prerecorded)": "1.2.1 Audio-only and Video-only(Prerecorded)",
  "1.2.2 Captions (Prerecorded)": "1.2.2 Captions(Prerecorded)",
  "1.2.3 Audio Description or Media Alternative (Prerecorded)": "1.2.3 Audio Description or Media Alternative (Prerecorded)",
  "1.2.4 Captions (Live)": "1.2.4 Captions (Live)",
  "1.2.5 Audio Description (Prerecorded)": "1.2.5 Audio Description (Prerecorded)",
  "1.3.2 Meaningful Sequence": "1.3.2 Meaningful Sequences",
  "1.3.3 Sensory Characteristics": "1.3.3 Sensory Characteristics",
  "1.3.4 Orientation": "1.3.4 Orientation",
  "1.3.5 Identify Input Purpose": "1.3.5 Identity Input Purpose",
  "1.4.1 Use of Color": "1.4.1 Use Of Color",
  "1.4.2 Audio Control": "1.4.2 Audio Control",
  "1.4.3 Contrast (Minimum)": "1.4.3 Contrast(Minimum)",
  "1.4.4 Resize text": "1.4.4 Resize Text",
  "1.4.5 Images of Text": "1.4.5 Images Of Text",
  "1.4.10 Reflow": "1.4.10 Reflow",
  "1.4.11 Non-text Contrast": "1.4.11 Non-Text Contrast",
  "1.4.12 Text Spacing": "1.4.12 Text Spacing",
  "1.4.13 Content on Hover or Focus": "1.4.13 Content On Hover Or Focus",
  "2.1.1 Keyboard": "2.1.1 Keyboard",
  "2.1.2 No Keyboard Trap": "2.1.2 No Keyboard Trap",
  "2.1.4 Character Key Shortcuts": "2.1.3 Character Key Shortcuts",
  "2.2.1 Timing Adjustable": "2.2.1 Timing Adjustable",
  "2.2.2 Pause, Stop, Hide": "2.2.2 Pause, Stop, Hide",
  "2.3.1 Three Flashes or Below Threshold": "2.3.1 Three Flashes Or Below Threshold",
  "2.4.1 Bypass Blocks": "2.4.1 Bypass Blocks",
  "2.4.4 Link Purpose (In Context)": "2.4.4 Link Purpose (In Context)",
  "2.4.5 Multiple Ways": "2.4.5 Multiple Ways",
  "2.4.6 Headings and Labels": "2.4.6 Headings And Labels",
  "2.4.7 Focus Visible": "2.4.7 Focus Visible",
  "2.5.1 Pointer Gestures": "2.5.1 Pointer Gestures",
  "2.5.2 Pointer Cancellation": "2.5.2 Pointer Cancellation",
  "2.5.3 Label in Name":"2.5.3 Label In  Name",
  "2.5.4 Motion Actuation": "2.5.4 Motion Actuation",
  "3.2.1 On Focus": "3.2.1 On Focus",
  "3.2.2 On Input":"3.2.2 On Input",
  "3.2.3 Consistent Navigation": "3.2.3 Consistent Navigation",
  "3.2.4 Consistent Identification": "3.2.4 Consistent Identification",
  "3.3.1 Error Identification": "3.3.1 Error Identification",
  "3.3.2 Labels or Instructstions": "3.3.2 Labels Or Instructions",
  "3.3.3 Error Suggestion": "3.3.3 Error Suggestion",
  "3.3.4 Error Prevention (Legal, Financial, Data)": "3.3.4 Error Prevention (Legal, Financial, Data) (AA)",
  "4.1.3 Status Messages": "4.1.3 Status Messages",
  "4.1.2 (PAC Not Available)": "4.1.2 Me, Role, Value",
  "Validity of document title": "PREP_NA",
  "Artifacted content on page body": "PREP_NA",
  "Tagged text consists of only whitespace": "PREP_NA",
  "Tagged content exists outside of the page boundaries": "PREP_NA",
  "Presence of headings": "PREP_NA",
  "Presence of bookmarks (document outline) if there are headings": "PREP_NA",
  "\"TOCI\" elements contain \"Link\" elements": "PREP_NA",
  "\"TOCI\" elements are correctly linked to headings": "PREP_NA",
  "Validity of alternative texts": "PREP_NA",
  "Alternative text on text elements": "PREP_NA",
  "Completeness of \"Link\" elements": "Completeness Of Links",
  "Formal correctness of \"LI\" elements": "PREP_NA",
  "Completeness of \"Table\" elements": "PREP_NA",
  "\"Note\" elements are referenced": "PREP_NA",
  "\"Note\" elements contain \"Lbl\" elements": "PREP_NA",
  "\"P\" elements contain \"Note\" elements": "PREP_NA",
   "accessibility permission flag": "pac_NA",
  "bookmarks": "pac_NA",
  "image only pdf": "pac_NA",
  "logical reading order": "pac_NA",
  "colour contrast": "pac_NA",
  "navigation links": "pac_NA",
  "screen flicker": "pac_NA",
  "scripts": "pac_NA",
  "timed responses": "pac_NA",
   "tagged multimedia": "pac_NA",
   "hides annotation": "pac_NA"
  
    }

    # Normalize JSON keys
    normalized_mapping = {
        k.strip().lower(): v for k, v in check_mapping.items()
    }

    # ======================================================
    # AUTO WIDTH FUNCTION
    # ======================================================
    def auto_adjust_column_width(file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

        wb.save(file_path)

    # ======================================================
    # COUNT FUNCTION
    # ======================================================
  
#     normalized_mapping = {
#     k.strip().lower(): v
#     for k, v in MAPPING.items()
# }
    def get_check_counts(df):

        df.columns = df.columns.str.strip()
        df.columns = df.columns.str.replace("  ", " ", regex=False)

        remove_cols = ["File Name", "File ID", "FileName", "Source ID", "SourceID"]
        df = df.drop(columns=[c for c in remove_cols if c in df.columns])

        result = {}

        for col in df.columns:

            clean_col = col.strip()
            key = clean_col.lower()

            if df[col].isna().all():
                continue

            values = df[col].astype(str).str.strip().str.lower()

            passed = values.isin(["passed", "pass"]).sum()
            failed = values.isin(["failed", "fail"]).sum()
            warning = values.isin(["warning", "warn"]).sum()
            skipped = values.isin(["skipped", "skip"]).sum()

            na_count = len(values) - (passed + failed + warning + skipped)

            result[key] = {
                "OriginalName": clean_col,
                "Passed": passed,
                "Failed": failed,
                "Warning": warning,
                "Skipped": skipped,
                "NA": na_count
            }

        return result

    # ======================================================
    # READ FILES
    # ======================================================
    pac_df = pd.read_excel(pac_file_path)
    prep_df = pd.read_excel(prep_file_path)

    pac_counts = get_check_counts(pac_df)
    prep_counts = get_check_counts(prep_df)

    final_rows = []

    # ======================================================
    # 1️⃣ LOOP THROUGH PAC (MASTER LEFT SIDE)
    # ======================================================
    for pac_key, pac_values in pac_counts.items():

        row = {}

        # LEFT SIDE (PAC)
        row["pac_CheckName"] = pac_values["OriginalName"]
        row["pac_Passed"] = pac_values["Passed"]
        row["pac_Failed"] = pac_values["Failed"]
        row["pac_Warning"] = pac_values["Warning"]
        row["pac_Skipped"] = pac_values["Skipped"]
        row["pac_na"] = pac_values["NA"]

        row[" "] = ""

        # RIGHT SIDE (PREP)
        mapped_value = normalized_mapping.get(pac_key)

        if mapped_value is None or mapped_value == "PREP_NA":
            row["prep_CheckName"] = ""
            row["prep_Passed"] = ""
            row["prep_Failed"] = ""
            row["prep_Warning"] = ""
            row["prep_Skipped"] = ""
            row["prep_na"] = ""

        elif mapped_value == "pac_NA":
            row["prep_CheckName"] = ""
            row["prep_Passed"] = ""
            row["prep_Failed"] = ""
            row["prep_Warning"] = ""
            row["prep_Skipped"] = ""
            row["prep_na"] = ""

        else:
            prep_key = mapped_value.strip().lower()

            if prep_key in prep_counts:
                prep_val = prep_counts[prep_key]
                row["prep_CheckName"] = prep_val["OriginalName"]
                row["prep_Passed"] = prep_val["Passed"]
                row["prep_Failed"] = prep_val["Failed"]
                row["prep_Warning"] = prep_val["Warning"]
                row["prep_Skipped"] = prep_val["Skipped"]
                row["prep_na"] = prep_val["NA"]
            else:
                row["prep_CheckName"] = mapped_value
                row["prep_Passed"] = ""
                row["prep_Failed"] = ""
                row["prep_Warning"] = ""
                row["prep_Skipped"] = ""
                row["prep_na"] = ""

        final_rows.append(row)

    # ======================================================
    # 2️⃣ ADD pac_NA CHECKS AT BOTTOM
    # ======================================================
    for json_key, mapped_value in normalized_mapping.items():

        if mapped_value == "pac_NA" and json_key in prep_counts:

            prep_val = prep_counts[json_key]

            row = {}

            # PAC side blank
            row["pac_CheckName"] = ""
            row["pac_Passed"] = ""
            row["pac_Failed"] = ""
            row["pac_Warning"] = ""
            row["pac_Skipped"] = ""
            row["pac_na"] = ""

            row[" "] = ""

            # PREP side filled
            row["prep_CheckName"] = prep_val["OriginalName"]
            row["prep_Passed"] = prep_val["Passed"]
            row["prep_Failed"] = prep_val["Failed"]
            row["prep_Warning"] = prep_val["Warning"]
            row["prep_Skipped"] = prep_val["Skipped"]
            row["prep_na"] = prep_val["NA"]

            final_rows.append(row)

    # ======================================================
    # BUILD FINAL EXCEL
    # ======================================================
    final_df = pd.DataFrame(final_rows)

    final_df.columns = [
    "pac_CheckName",
    "pac_Passed",
    "pac_Failed",
    "pac_Warning",
    "pac_Skipped",
    "pac_na",
    "",
    "prep_CheckName",
    "prep_Passed",
    "prep_Failed",
    "prep_Warning",
    "prep_Skipped",
    "prep_na"
]
  
    
    final_df.to_excel(output_path, index=False, header=False, startrow=1)

          # Fix duplicate column names in Excel
    wb = load_workbook(output_path)
    ws = wb.active

    headers = [
    "CheckName", "Passed", "Failed", "Warning", "Skipped", "pac_na",
    "", 
    "CheckName", "Passed", "Failed", "Warning", "Skipped", "prep_na"
    ]

    for col, header in enumerate(headers, start=1):
       ws.cell(row=1, column=col).value = header

    wb.save(output_path)
    auto_adjust_column_width(output_path)

    print("✅ Final PAC vs PREP report generated successfully.")
    return output_path

#step 7 slack message
def send_slack_summary(version_summary_file,result):

    print("🔄 Sending Slack Summary...")

    load_dotenv()
    SLACK_TOKEN = os.getenv("SLACK_TOKEN")
    CHANNEL_ID  = os.getenv("CHANNEL_ID")

    client = WebClient(token=SLACK_TOKEN)

    summary_df = pd.read_excel(version_summary_file)
    row = summary_df.iloc[-1]

    # ── Totals & percentages ──────────────────────────────────────────────────
    agreement_total = (result['both_skipped'] + result['both_passed']
                       + result['both_failed'])
    mismatch_total  = (result['prep_passed_pac_failed'] + result['prep_failed_pac_passed'] +
                       result['prep_skipped_pac_passed'] + result['prep_skipped_pac_failed'] +
                       result['prep_passed_pac_skipped'] + result['prep_failed_pac_skipped'])
    grand_total = agreement_total + mismatch_total

    def pct(val):
        return f"{val / grand_total * 100:.1f}%" if grand_total else "0.0%"

    def trow(icon, label, count):
        return f"  {icon} {label:<34} {count:>6}   {pct(count):>6}"

    sep   = "─" * 58
    table = (
        f"```\n"
        f"  {'Check':<35} {'Count':>6}   {'% Total':>7}\n"
        f"{sep}\n"
        f"  ✅ AGREEMENT — {agreement_total:,} ({pct(agreement_total)})\n"
        f"{sep}\n"
        f"{trow('✅', 'PAC & PREP Passed',           result['both_passed'])}\n"
        f"{trow('✅', 'PAC & PREP Failed',           result['both_failed'])}\n"
        f"{trow('✅', 'PAC & PREP Skipped',          result['both_skipped'])}\n"
        f"{sep}\n"
        f"  ❌ MISMATCH — {mismatch_total:,} ({pct(mismatch_total)})\n"
        f"{sep}\n"
        f"{trow('❌', 'PREP Passed  / PAC Failed',   result['prep_passed_pac_failed'])}\n"
        f"{trow('❌', 'PREP Failed  / PAC Passed',   result['prep_failed_pac_passed'])}\n"
        f"{trow('❌', 'PREP Skipped / PAC Passed',   result['prep_skipped_pac_passed'])}\n"
        f"{trow('❌', 'PREP Skipped / PAC Failed',   result['prep_skipped_pac_failed'])}\n"
        f"{trow('❌', 'PREP Passed  / PAC Skipped',  result['prep_passed_pac_skipped'])}\n"
        f"{trow('❌', 'PREP Failed  / PAC Skipped',  result['prep_failed_pac_skipped'])}\n"
        f"{sep}\n"
        f"  📌 TOTAL INSTANCES{'':>17} {grand_total:>6}   100.0%\n"
        f"```"
    )

    # ── Blocks ────────────────────────────────────────────────────────────────
    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": "📊 Accessibility Report Summary — PAC vs PREP"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "fields": [
                {"type": "mrkdwn", "text": f"*Version:*\n{row['Version']}"},
                {"type": "mrkdwn", "text": f"*Date:*\n{row['Date']}"},
                {"type": "mrkdwn", "text": f"*Server Name:*\n{row['Server_Name']}"},
                {"type": "mrkdwn", "text": f"*Total Files:*\n{row['Total_Files']}"},
            ]
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": (
                    f"*Overall Alignment*\n"
                    f"✅  *Agreement* (tools match): `{agreement_total:,}` — *{pct(agreement_total)}*\n"
                    f"❌  *Mismatch*  (tools differ): `{mismatch_total:,}` — *{pct(mismatch_total)}*"
                )
            }
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"*Full Breakdown*\n{table}"}
        },
    ]

    client.chat_postMessage(
        channel=CHANNEL_ID,
        blocks=blocks,
        text="Accessibility Report Summary"
    )

    print("✅ Slack summary sent successfully!")
    print("✅ Slack report sent successfully")


# =====================================================
# 🚀 MASTER FLOW here all the functions called
# =====================================================
def run_full_automation():

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    VERSION_FOLDER = os.path.join(BASE_DIR, version)
    os.makedirs(VERSION_FOLDER, exist_ok=True)

    pac_file = os.path.join(
        VERSION_FOLDER,
        "pac_results",
        f"PAC_Final_Summary_{version}.xlsx"
    )

    prep_file = os.path.join(
        VERSION_FOLDER,
        "prep_results",
        f"prep_final_summary_{version}.xlsx"
    )


    final_report_file = os.path.join(
        VERSION_FOLDER,
        f"PrepPac_Comparison_Report_{version}.xlsx"
        
    )

    output_file=os.path.join(
        VERSION_FOLDER, f"PAC_PREP_Final_{version}.xlsx"
        )

    status_file = os.path.join(
        VERSION_FOLDER,
  
        f"final_accessibility_status_only_{version}.xlsx"
    )

    version_summary_file = os.path.join(
        VERSION_FOLDER,
        f"version_summary_report_{version}.xlsx"
    
    )

    splitup_report = os.path.join(
    VERSION_FOLDER,
    f"splitup_report_{version}.xlsx"
   )
    
 
    final_report = generate_accessibility_report( pac_file,prep_file,final_report_file)

    result = generate_comparison_report(pac_file,prep_file,output_file )
    
    status_output= generate_status_file(final_report, status_file)

    version_summary_output=generate_version_summary(status_output, version_summary_file)

    column_summary_output=generate_final_column_summary(final_report)

    prep_pac_splitup=generate_pac_prep_report(pac_file, prep_file,splitup_report)

    send_slack_summary(version_summary_output,  result )
    

    print("\n🚀 FULL AUTOMATION COMPLETED SUCCESSFULLY 🚀")







def cleanup_folders(version_folder):
    folders_to_delete = ["working","processed"]

    for folder in folders_to_delete:
        folder_path = os.path.join(version_folder, folder)

        if os.path.exists(folder_path):
            try:
                # wait a little to release file locks
                time.sleep(1)

                shutil.rmtree(folder_path, ignore_errors=True)

                if not os.path.exists(folder_path):
                    print(f"🧹 Deleted folder: {folder}")
                else:
                    print(f"⚠️ Could not delete {folder} (possibly locked)")
            except Exception as e:
                print(f"⚠️ Error deleting {folder}: {e}")

def combine_reports(version_folder, version):

    combined_file = os.path.join(
        version_folder,
        f"Final_Accessibility_Report_{version}.xlsx"
    )

    files = {
        "Comparison_Report": f"PrepPac_Comparison_Report_{version}.xlsx",
        "Status": f"final_accessibility_status_only_{version}.xlsx",
        "Version_Summary": f"version_summary_report_{version}.xlsx",
        "Column_Summary": f"final_summary_{version}.xlsx",
        "Splitup": f"splitup_report_{version}.xlsx"
    }

    with pd.ExcelWriter(combined_file, engine="openpyxl") as writer:

        for sheet_name, file_name in files.items():

            file_path = os.path.join(version_folder, file_name)

            if os.path.exists(file_path):

                # Ensure source workbook handle is released on Windows before cleanup.
                with pd.ExcelFile(file_path) as excel_file:
                    for sheet in excel_file.sheet_names:

                        df = pd.read_excel(file_path, sheet_name=sheet)

                        new_sheet_name = f"{sheet_name}_{sheet}"[:31]

                        df.to_excel(writer, sheet_name=new_sheet_name, index=False)

    # Auto resize columns
    wb = load_workbook(combined_file)

    for ws in wb.worksheets:
        for column_cells in ws.columns:

            max_length = 0
            col_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(combined_file)

    print(f"✅ Combined Excel created successfully: {combined_file}")


# ============================================================

def cleanup_files(version_folder, version):

    import gc
    gc.collect()
    time.sleep(2)
    for file in os.listdir(version_folder):

        if file.endswith(".xlsx") and file not in [
            f"Final_Accessibility_Report_{version}.xlsx",
            f"PAC_PREP_Final_{version}.xlsx"
        ]:

            path = os.path.join(version_folder, file)

            for _ in range(8):
                try:
                    os.remove(path)
                    print(f"🗑 Deleted: {file}")
                    break
                except PermissionError:
                    print(f"⚠️ File locked, retrying: {file}")
                    time.sleep(5)



if __name__ == "__main__":

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    VERSION_FOLDER = os.path.join(BASE_DIR, version)

    run_full_automation()

    time.sleep(5)

    # Combine accessibility reports
    combine_reports(VERSION_FOLDER, version)

    cleanup_files(VERSION_FOLDER, version)

    # Delete working folders
    cleanup_folders(VERSION_FOLDER)
   
 
